"""
portfolio_backtest.py — combine multiple fixed-lot backtests into a
single portfolio equity view.

Fully standalone — does NOT depend on basket_analysis.py or any other
local module. Only uses the Python standard library.

Each strategy contributes its raw dollar P&L curve (no normalization,
no percentage returns). The combined view sums them on a unified daily
timeline, so it shows what the combined account would have looked like
if every strategy traded side-by-side with fixed lots.

A per-strategy scaling factor multiplies BOTH realized P&L AND lot
sizes, so floating-loss drawdowns scale correctly alongside closed
profits. Default scale is 1.0.

Supported backtest formats (auto-detected):
  * MT4 live account statement (UTF-8 HTML, 14-column trade rows)
  * MT4 strategy tester report (UTF-8 HTML, separate open/close rows)
  * MT5 strategy tester report (UTF-16 HTML, Deals table with in/out)

Bar data must be CSV with unix_ts,open,high,low,close columns. Any
timeframe works (M1/M5/M15/H1/...) since bars are only used to
sample the equity curve between trade events.

Usage:
    python portfolio_backtest.py ^
        --strategy "EURUSD|eurusd_bt.html|eurusd_m5.csv|1.0|2" ^
        --strategy "USDCAD|usdcad_bt.html|usdcad_m5.csv|1.0|2" ^
        --out-dir C:\\Trading\\Portfolio

Strategy string is pipe-delimited:
    SYMBOL|BACKTEST_PATH|BARS_CSV|SCALE|BROKER_GMT
Scale and GMT are optional (defaults 1.0 and 2).
"""

import argparse
import csv
import html as html_lib
import html.parser as html_parser
import json
import os
import re
import sys
import time
from datetime import datetime, timedelta, timezone
from typing import Dict, List, Optional, Tuple


# ────────────────────────────────────────────────────────────────────────────
# File reading with encoding auto-detection
# ────────────────────────────────────────────────────────────────────────────
def read_text_file(path: str) -> str:
    """Read a text file as str, handling UTF-8 and UTF-16 (BOM-sniffed)."""
    with open(path, "rb") as fh:
        raw = fh.read()
    if raw[:2] in (b"\xff\xfe", b"\xfe\xff"):
        return raw.decode("utf-16")
    if raw[:3] == b"\xef\xbb\xbf":
        return raw[3:].decode("utf-8", errors="replace")
    # Heuristic fallback: MT5 reports are sometimes UTF-16 without BOM
    if raw[:200].count(b"\x00") > 50:
        try:
            return raw.decode("utf-16")
        except UnicodeDecodeError:
            pass
    return raw.decode("utf-8", errors="replace")


# ────────────────────────────────────────────────────────────────────────────
# HTML table row extraction
# ────────────────────────────────────────────────────────────────────────────
class _RowParser(html_parser.HTMLParser):
    """Collects plain-text cells from every <tr>."""

    def __init__(self):
        super().__init__()
        self.rows: List[List[str]] = []
        self._row: Optional[List[str]] = None
        self._cell: Optional[List[str]] = None

    def handle_starttag(self, tag, attrs):
        tag = tag.lower()
        if tag == "tr":
            if self._row is not None:
                self.rows.append(self._row)
            self._row = []
            self._cell = None
        elif tag in ("td", "th"):
            if self._cell is not None and self._row is not None:
                self._row.append("".join(self._cell).strip())
            self._cell = []

    def handle_endtag(self, tag):
        tag = tag.lower()
        if tag in ("td", "th"):
            if self._cell is not None and self._row is not None:
                self._row.append("".join(self._cell).strip())
                self._cell = None
        elif tag == "tr":
            if self._row is not None:
                if self._cell is not None:
                    self._row.append("".join(self._cell).strip())
                    self._cell = None
                self.rows.append(self._row)
                self._row = None

    def handle_data(self, data):
        if self._cell is not None:
            self._cell.append(data)

    def close(self):
        if self._row is not None:
            if self._cell is not None:
                self._row.append("".join(self._cell).strip())
            self.rows.append(self._row)
            self._row = None
            self._cell = None
        super().close()


def _extract_rows(html_text: str) -> List[List[str]]:
    parser = _RowParser()
    parser.feed(html_text)
    parser.close()
    return [[c.replace("\xa0", " ").strip() for c in row] for row in parser.rows]


# ────────────────────────────────────────────────────────────────────────────
# Format detection
# ────────────────────────────────────────────────────────────────────────────
def _detect_format(html_text: str, rows: List[List[str]]) -> str:
    """Return one of: 'mt5_tester', 'mt4_tester', 'mt4_live'."""
    if re.search(r">\s*Deals\s*<", html_text, re.IGNORECASE) and re.search(r"Direction", html_text, re.IGNORECASE):
        for row in rows:
            if len(row) >= 13 and row[4].strip().lower() in ("in", "out"):
                return "mt5_tester"

    has_tester_open = any(
        len(row) >= 6 and row[2].strip().lower() in ("buy", "sell")
        for row in rows
    )
    has_tester_close = any(
        len(row) >= 6 and _is_mt4_tester_close_type(row[2])
        for row in rows
    )
    has_tester_header = any(
        {"#", "time", "type", "order", "size", "price", "profit", "balance"}.issubset(
            {_normalize_header_name(c) for c in row if c.strip()}
        )
        for row in rows
    )
    if "strategy tester report" in html_text.lower() and (has_tester_header or has_tester_open or has_tester_close):
        return "mt4_tester"
    if has_tester_header or (has_tester_open and has_tester_close):
        return "mt4_tester"

    has_live_trade = any(
        len(row) >= 13 and row[2].strip().lower() in ("buy", "sell")
        for row in rows
    )
    if has_live_trade:
        return "mt4_live"

    raise ValueError(
        "Could not detect backtest format. Expected MT4 live statement, "
        "MT4 strategy tester, or MT5 strategy tester report."
    )


# ────────────────────────────────────────────────────────────────────────────
# Trade parsers
# ────────────────────────────────────────────────────────────────────────────
def _parse_dt(s: str, offset: timezone) -> int:
    s = s.strip()
    for fmt in ("%Y.%m.%d %H:%M:%S", "%Y.%m.%d %H:%M"):
        try:
            return int(datetime.strptime(s, fmt).replace(tzinfo=offset).timestamp())
        except ValueError:
            continue
    raise ValueError(f"Unrecognised datetime: {s!r}")


def _parse_num(s: str) -> float:
    """Parse a number that may use space as thousands separator."""
    return float(s.replace(" ", "").replace(",", ""))


def _normalize_header_name(value: str) -> str:
    return re.sub(r"\s+", " ", (value or "").strip().lower()).rstrip(":")


def _is_mt4_tester_close_type(value: str) -> bool:
    token = re.sub(r"\s+", " ", (value or "").strip().lower())
    if not token:
        return False
    if token in (
        "close",
        "t/p",
        "s/l",
        "tp",
        "sl",
        "stop loss",
        "take profit",
        "close at stop",
        "close at limit",
    ):
        return True
    return token.startswith("close")


def _parse_mt4_live(rows, offset, symbol_filter):
    trades = []
    for row in rows:
        if len(row) != 14:
            continue
        ttype = row[2].strip().lower()
        if ttype not in ("buy", "sell"):
            continue
        if symbol_filter and symbol_filter.lower() not in row[4].strip().lower():
            continue
        try:
            trades.append({
                "type": ttype,
                "ts": _parse_dt(row[1], offset),
                "close_ts": _parse_dt(row[8], offset),
                "price": float(row[5]),
                "close_price": float(row[9]),
                "lots": float(row[3]),
                "profit": _parse_num(row[13]),
                "commission": _parse_num(row[10]),
                "swap": _parse_num(row[12]),
            })
        except (ValueError, IndexError):
            continue
    return trades


def _parse_mt4_tester(rows, offset):
    opens: Dict[int, Dict] = {}
    trades = []
    ignored_types = {"modify", "delete", "balance", "credit"}

    for row in rows:
        if len(row) < 6:
            continue

        ttype = row[2].strip().lower()
        if ttype in ("buy", "sell"):
            try:
                order = int(float(row[3]))
                opens[order] = {
                    "type": ttype,
                    "ts": _parse_dt(row[1], offset),
                    "price": _parse_num(row[5]),
                    "lots": _parse_num(row[4]),
                }
            except (ValueError, IndexError):
                pass
            continue

        if ttype in ignored_types or not _is_mt4_tester_close_type(ttype):
            continue

        try:
            order = int(float(row[3]))
            if order not in opens:
                continue
            if len(row) <= 8 or not row[8].strip():
                continue

            o = opens.pop(order)
            trades.append({
                "type": o["type"],
                "ts": o["ts"],
                "close_ts": _parse_dt(row[1], offset),
                "price": o["price"],
                "close_price": _parse_num(row[5]),
                "lots": o["lots"],
                "profit": _parse_num(row[8]),
                "commission": 0.0,
                "swap": 0.0,
            })
        except (ValueError, IndexError):
            continue
    return trades


def _parse_mt5_tester(rows, offset, symbol_filter):
    """
    MT5 Deals table columns:
      0 Time | 1 Deal | 2 Symbol | 3 Type | 4 Direction | 5 Volume
      6 Price | 7 Order | 8 Commission | 9 Swap | 10 Profit
      11 Balance | 12 Comment

    Each position yields an 'in' deal (opens) and an 'out' deal (closes).
    FIFO-match each 'out' to the oldest 'in' of the OPPOSITE type. This
    is approximate for basket strategies where many positions close at
    once, but the aggregate P&L and timestamps are correct — the Profit
    on each 'out' deal is already the right dollar amount regardless of
    which 'in' we pair it with.
    """
    open_fifo: List[Dict] = []
    trades: List[Dict] = []

    for row in rows:
        if len(row) < 13:
            continue
        direction = row[4].strip().lower()
        if direction not in ("in", "out"):
            continue
        ttype = row[3].strip().lower()
        if ttype not in ("buy", "sell"):
            continue

        symbol = row[2].strip()
        if symbol_filter and symbol_filter.lower() not in symbol.lower():
            continue

        try:
            ts = _parse_dt(row[0], offset)
            volume = _parse_num(row[5])
            price = _parse_num(row[6])
            commission = _parse_num(row[8]) if row[8] else 0.0
            swap = _parse_num(row[9]) if row[9] else 0.0
            profit = _parse_num(row[10]) if row[10] else 0.0
        except (ValueError, IndexError):
            continue

        if direction == "in":
            open_fifo.append({
                "type": ttype,
                "ts": ts,
                "price": price,
                "lots": volume,
                "open_commission": commission,
                "open_swap": swap,
            })
        else:
            target_type = "buy" if ttype == "sell" else "sell"
            match_idx = next(
                (i for i, o in enumerate(open_fifo) if o["type"] == target_type),
                None
            )
            if match_idx is None:
                continue
            o = open_fifo.pop(match_idx)
            trades.append({
                "type": o["type"],
                "ts": o["ts"],
                "close_ts": ts,
                "price": o["price"],
                "close_price": price,
                "lots": o["lots"],
                "profit": profit,
                "commission": commission + o["open_commission"],
                "swap": swap + o["open_swap"],
            })

    return trades


def parse_backtest(path: str, broker_gmt: int,
                    symbol_filter: Optional[str]) -> Tuple[List[Dict], str]:
    """Auto-detect format and return (trades, format_name)."""
    html_text = read_text_file(path)
    rows = _extract_rows(html_text)
    fmt = _detect_format(html_text, rows)
    offset = timezone(timedelta(hours=broker_gmt))

    if fmt == "mt5_tester":
        trades = _parse_mt5_tester(rows, offset, symbol_filter)
    elif fmt == "mt4_tester":
        trades = _parse_mt4_tester(rows, offset)
    else:
        trades = _parse_mt4_live(rows, offset, symbol_filter)

    trades.sort(key=lambda t: t["ts"])
    return trades, fmt


# ────────────────────────────────────────────────────────────────────────────
# Bar CSV loading
# ────────────────────────────────────────────────────────────────────────────
def load_bars(path: str) -> List[Dict]:
    """Load OHLC bars from CSV. Expected columns: unix_ts, o, h, l, c."""
    bars: List[Dict] = []
    with open(path, "r", encoding="utf-8") as fh:
        reader = csv.reader(fh)
        for row in reader:
            if len(row) < 5:
                continue
            try:
                bars.append({
                    "ts": int(float(row[0])),
                    "o": float(row[1]),
                    "h": float(row[2]),
                    "l": float(row[3]),
                    "c": float(row[4]),
                })
            except (ValueError, IndexError):
                continue
    bars.sort(key=lambda b: b["ts"])
    return bars


def load_ticks(path: str, tick_gmt: int,
               min_ts: Optional[int] = None,
               max_ts: Optional[int] = None,
               progress_every: int = 1_000_000) -> List[Dict]:
    """Load ticks from CSV. Expected layout: timestamp,bid,ask in first 3 cols."""
    ticks: List[Dict] = []
    offset = timezone(timedelta(hours=tick_gmt))
    processed = 0
    skipped_outside_window = 0

    with open(path, "r", encoding="utf-8") as fh:
        reader = csv.reader(fh)
        for row_num, row in enumerate(reader):
            processed += 1
            if progress_every > 0 and processed % progress_every == 0:
                print(f"  Tick load progress: {processed:,} rows scanned, {len(ticks):,} kept", flush=True)

            if len(row) < 3:
                continue
            try:
                if row_num == 0 and row[0].lower() in ("time", "date", "datetime"):
                    continue

                ts_str = row[0].strip()
                for fmt in ("%d.%m.%Y %H:%M:%S.%f", "%d.%m.%Y %H:%M:%S",
                            "%Y-%m-%d %H:%M:%S", "%Y.%m.%d %H:%M:%S"):
                    try:
                        dt = datetime.strptime(ts_str, fmt)
                        break
                    except ValueError:
                        continue
                else:
                    continue

                ts = int(dt.replace(tzinfo=offset).timestamp())
                if min_ts is not None and ts < min_ts:
                    skipped_outside_window += 1
                    continue
                if max_ts is not None and ts > max_ts:
                    break

                bid = float(row[1])
                ask = float(row[2])
                ticks.append({"ts": ts, "bid": bid, "ask": ask})
            except (ValueError, IndexError):
                continue

    if ticks and any(ticks[i]["ts"] > ticks[i + 1]["ts"] for i in range(min(len(ticks) - 1, 1000))):
        ticks.sort(key=lambda t: t["ts"])

    print(
        f"  Tick load complete: {processed:,} rows scanned, {len(ticks):,} kept"
        f" ({skipped_outside_window:,} before window)",
        flush=True,
    )
    return ticks


# ────────────────────────────────────────────────────────────────────────────
# Equity curve
# ────────────────────────────────────────────────────────────────────────────
def _pip_size(price: float) -> float:
    return 0.01 if price > 20 else 0.0001


def _trade_mtm(ttype: str, open_price: float, current_price: float,
                lots: float, pip_size: float) -> float:
    """Approximate mark-to-market P&L at $10/pip per 1.0 lot."""
    pip_move = (current_price - open_price) / pip_size
    if ttype == "sell":
        pip_move = -pip_move
    return pip_move * lots * 10.0


def build_equity_curve(trades: List[Dict], bars: List[Dict],
                        sample_every: int = 15) -> List[Dict]:
    curves: List[Dict] = []
    realised = 0.0

    trades_sorted = sorted(trades, key=lambda t: t["ts"])
    open_idx = 0
    active: List[Dict] = []

    for bi in range(0, len(bars), sample_every):
        bar = bars[bi]
        bar_ts = bar["ts"]
        bar_c = bar["c"]

        while open_idx < len(trades_sorted) and trades_sorted[open_idx]["ts"] <= bar_ts:
            active.append(trades_sorted[open_idx])
            open_idx += 1

        still: List[Dict] = []
        for t in active:
            if t["close_ts"] <= bar_ts:
                realised += t["profit"] + t["commission"] + t["swap"]
            else:
                still.append(t)
        active = still

        unreal = 0.0
        for t in active:
            ps = _pip_size(t["price"])
            unreal += _trade_mtm(t["type"], t["price"], bar_c, t["lots"], ps)

        curves.append({
            "ts": bar_ts,
            "bal": round(realised, 2),
            "eq": round(realised + unreal, 2),
        })

    while open_idx < len(trades_sorted):
        active.append(trades_sorted[open_idx])
        open_idx += 1
    for t in active:
        realised += t["profit"] + t["commission"] + t["swap"]

    if trades_sorted:
        final_ts = max(
            curves[-1]["ts"] if curves else 0,
            trades_sorted[-1]["close_ts"],
        )
        if curves and curves[-1]["bal"] != round(realised, 2):
            curves.append({
                "ts": final_ts,
                "bal": round(realised, 2),
                "eq": round(realised, 2),
            })

    return curves


def build_equity_curve_from_ticks(trades: List[Dict], ticks: List[Dict],
                                  sample_every: int = 100) -> List[Dict]:
    """Build equity curve from tick data by computing mark-to-market P&L."""
    curves: List[Dict] = []
    realised = 0.0

    trades_sorted = sorted(trades, key=lambda t: t["ts"])
    tick_idx = 0
    active: List[Dict] = []

    for ti in range(0, len(ticks), sample_every):
        tick = ticks[ti]
        tick_ts = tick["ts"]
        mid = (tick["bid"] + tick["ask"]) / 2.0

        while tick_idx < len(trades_sorted) and trades_sorted[tick_idx]["ts"] <= tick_ts:
            active.append(trades_sorted[tick_idx])
            tick_idx += 1

        still: List[Dict] = []
        for t in active:
            if t["close_ts"] <= tick_ts:
                realised += t["profit"] + t["commission"] + t["swap"]
            else:
                still.append(t)
        active = still

        unreal = 0.0
        for t in active:
            ps = _pip_size(t["price"])
            unreal += _trade_mtm(t["type"], t["price"], mid, t["lots"], ps)

        curves.append({
            "ts": tick_ts,
            "bal": round(realised, 2),
            "eq": round(realised + unreal, 2),
        })

    while tick_idx < len(trades_sorted):
        active.append(trades_sorted[tick_idx])
        tick_idx += 1
    for t in active:
        realised += t["profit"] + t["commission"] + t["swap"]

    if trades_sorted:
        final_ts = max(
            curves[-1]["ts"] if curves else 0,
            trades_sorted[-1]["close_ts"],
        )
        if curves and curves[-1]["bal"] != round(realised, 2):
            curves.append({
                "ts": final_ts,
                "bal": round(realised, 2),
                "eq": round(realised, 2),
            })

    return curves


def curves_to_daily(curves: List[Dict]) -> Tuple[List[str], List[float], List[float]]:
    """Take the last sample per UTC calendar day."""
    by_day: Dict[str, Dict] = {}
    for c in curves:
        d = datetime.fromtimestamp(c["ts"], tz=timezone.utc).strftime("%Y-%m-%d")
        by_day[d] = c
    days = sorted(by_day)
    return days, [by_day[d]["bal"] for d in days], [by_day[d]["eq"] for d in days]


def filter_trades_to_recent_months(trades: List[Dict], months: float) -> List[Dict]:
    """Return only trades whose open timestamp falls within the most
    recent `months` of the final close in the backtest.

    Used when the user specifies --backtest-months to shorten a long
    backtest window to just the recent period (e.g. 'use only the last
    12 months of this 5-year backtest'). Trades that opened before the
    cutoff are dropped entirely; trades that opened after are kept in
    full regardless of when they closed.
    """
    if not trades or months is None or months <= 0:
        return trades
    last_close = max(t["close_ts"] for t in trades)
    window_seconds = int(months * 30.44 * 86400)
    cutoff = last_close - window_seconds
    return [t for t in trades if t["ts"] >= cutoff]


def max_drawdown(values: List[float]) -> Tuple[float, float, float]:
    """Return (peak, low, max_dd). Running peak starts at max(first, 0)."""
    if not values:
        return 0.0, 0.0, 0.0
    peak = max(values[0], 0.0)
    low = values[0]
    max_dd = 0.0
    running_peak = max(values[0], 0.0)
    for v in values:
        if v > running_peak:
            running_peak = v
        dd = running_peak - v
        if dd > max_dd:
            max_dd = dd
        if v > peak:
            peak = v
        if v < low:
            low = v
    return peak, low, max_dd


def months_between(date_strs: List[str]) -> float:
    """Return elapsed months between the first and last date string
    ('YYYY-MM-DD'). Uses 30.44 days/month average. Returns 0.0 for
    empty or single-day ranges."""
    if not date_strs or len(date_strs) < 2:
        return 0.0
    try:
        first = datetime.strptime(date_strs[0], "%Y-%m-%d")
        last = datetime.strptime(date_strs[-1], "%Y-%m-%d")
    except ValueError:
        return 0.0
    days = (last - first).days
    return days / 30.44 if days > 0 else 0.0


def max_open_positions(trades: List[Dict]) -> int:
    """Return the max number of simultaneously open trades.

    Close events are processed before open events at the same timestamp so
    back-to-back trades do not count as overlapping unless they truly do.
    """
    if not trades:
        return 0

    events: List[Tuple[int, int]] = []
    for t in trades:
        open_ts = int(t["ts"])
        close_ts = int(t["close_ts"])
        events.append((open_ts, 1))
        events.append((close_ts, 0))

    events.sort()
    current = 0
    peak = 0
    for _, event_type in events:
        if event_type == 0:
            current = max(0, current - 1)
        else:
            current += 1
            if current > peak:
                peak = current
    return peak


def max_open_positions_for_strategies(strategies: List[Dict]) -> int:
    """Return max simultaneous open trades across multiple strategies."""
    if not strategies:
        return 0

    events: List[Tuple[int, int]] = []
    for s in strategies:
        for open_ts, close_ts in s.get("trade_windows", []):
            events.append((int(open_ts), 1))
            events.append((int(close_ts), 0))

    if not events:
        return 0

    events.sort()
    current = 0
    peak = 0
    for _, event_type in events:
        if event_type == 0:
            current = max(0, current - 1)
        else:
            current += 1
            if current > peak:
                peak = current
    return peak


def format_elapsed(seconds: float) -> str:
    total = max(0, int(seconds))
    hours, rem = divmod(total, 3600)
    minutes, secs = divmod(rem, 60)
    if hours > 0:
        return f"{hours:d}:{minutes:02d}:{secs:02d}"
    return f"{minutes:02d}:{secs:02d}"


def compute_risk_metrics(net: float, max_dd: float, months: float,
                          account_size: float, dd_tolerance_pct: float
                          ) -> Dict:
    """Compute safety factor and monthly % estimate.

    Safety factor = (account × dd_tolerance) / max_dd
        Interpretation: the multiple of current size you could trade
        before the historical max DD would equal your tolerance limit.
        >1 = within budget, <1 = already exceeds budget.

    Monthly % = (net / account) / months
        Average monthly return as a fraction of the account over the
        backtest period.
    """
    allowable_dd = account_size * (dd_tolerance_pct / 100.0)
    if max_dd > 0:
        safety_factor = allowable_dd / max_dd
    else:
        safety_factor = float("inf")
    if months > 0 and account_size > 0:
        monthly_pct = (net / account_size) / months * 100.0
    else:
        monthly_pct = 0.0
    return {
        "allowable_dd": allowable_dd,
        "safety_factor": safety_factor,
        "monthly_pct": monthly_pct,
    }


# ────────────────────────────────────────────────────────────────────────────
# Strategy loading / combining / rendering
# ────────────────────────────────────────────────────────────────────────────
STRATEGY_COLORS = [
    "#378ADD", "#E89611", "#2E9E5A", "#B83D8A",
    "#7C4DFF", "#C94444", "#00897B", "#F4B400",
]
COMBINED_COLOR = "#111111"


def parse_strategy_arg(s: str) -> Dict:
    parts = [p.strip() for p in s.split("|")]
    if len(parts) < 2:
        raise ValueError(
            f"--strategy needs at least SYMBOL|BACKTEST (got: {s!r})"
        )
    symbol = parts[0]
    return {
        "symbol": symbol,
        "bt_path": parts[1],
        "bars_path": parts[2] if len(parts) >= 3 else "",
        "scale": float(parts[3]) if len(parts) >= 4 and parts[3] else 1.0,
        "broker_gmt": int(parts[4]) if len(parts) >= 5 and parts[4] else 2,
        # 6th field is optional:
        #   omitted              → filter defaults to display symbol
        #   non-empty value      → explicit filter override
        #   empty (trailing '|') → explicitly disable filtering
        "symbol_filter": (parts[5] if (len(parts) >= 6 and parts[5])
                           else (symbol if len(parts) < 6 else "")),
    }


def load_strategy(cfg: Dict, months_override: Optional[float] = None,
                  curve_source: str = "bars", tick_gmt: int = 2) -> Dict:
    print(f"\nLoading {cfg['symbol']} (scale={cfg['scale']:.2f}x)…")
    print(f"  Backtest: {cfg['bt_path']}")

    if not os.path.exists(cfg["bt_path"]):
        # Be helpful about the common .htm <-> .html mixup — try the
        # other extension automatically before giving up. Both formats
        # are used in the wild (MT4/MT5 sometimes save one, sometimes
        # the other, depending on version and export options).
        alt = None
        if cfg["bt_path"].lower().endswith(".html"):
            alt = cfg["bt_path"][:-1]  # .html -> .htm
        elif cfg["bt_path"].lower().endswith(".htm"):
            alt = cfg["bt_path"] + "l"  # .htm -> .html
        if alt and os.path.exists(alt):
            print(f"  (Using {alt} — the .htm/.html variant of the "
                  f"specified path)")
            cfg["bt_path"] = alt
        else:
            raise FileNotFoundError(f"Backtest file not found: {cfg['bt_path']}")

    trades, fmt = parse_backtest(cfg["bt_path"], cfg["broker_gmt"],
                                  cfg["symbol_filter"])
    print(f"  Detected format: {fmt}")
    if not trades:
        raise ValueError(
            f"{cfg['symbol']}: no trades parsed. Check --symbol filter "
            f"or file format."
        )

    total_trades_before_filter = len(trades)
    if months_override is not None and months_override > 0:
        trades = filter_trades_to_recent_months(trades, months_override)
        dropped = total_trades_before_filter - len(trades)
        if dropped > 0:
            print(f"  Filtered to most recent {months_override} months: "
                  f"kept {len(trades)}/{total_trades_before_filter} trades "
                  f"(dropped {dropped})")
        if not trades:
            raise ValueError(
                f"{cfg['symbol']}: no trades in the most recent "
                f"{months_override} months — the backtest may not span "
                f"that window."
            )

    from collections import Counter
    lot_counter = Counter(round(t["lots"], 4) for t in trades)
    base_lot = lot_counter.most_common(1)[0][0]

    scale = cfg["scale"]
    if scale != 1.0:
        for t in trades:
            t["profit"] *= scale
            t["commission"] *= scale
            t["swap"] *= scale
            t["lots"] *= scale

    if curve_source == "bars":
        print(f"  Bars:     {cfg.get('bars_path', '')}")
        if not cfg.get("bars_path") or not os.path.exists(cfg["bars_path"]):
            raise FileNotFoundError(f"Bars file not found: {cfg.get('bars_path', '')}")
        bars = load_bars(cfg["bars_path"])
        if not bars:
            raise ValueError(f"{cfg['symbol']}: no bars loaded from {cfg['bars_path']}")
        curves = build_equity_curve(trades, bars)
    elif curve_source == "ticks":
        tick_path = cfg.get("tick_path", "")
        print(f"  Ticks:    {tick_path}")
        if not tick_path or not os.path.exists(tick_path):
            raise FileNotFoundError(f"Tick file not found: {tick_path}")
        min_ts = min(t["ts"] for t in trades)
        max_ts = max(t["close_ts"] for t in trades)
        ticks = load_ticks(tick_path, tick_gmt, min_ts=min_ts, max_ts=max_ts)
        if not ticks:
            raise ValueError(f"{cfg['symbol']}: no ticks loaded from {tick_path}")
        curves = build_equity_curve_from_ticks(trades, ticks)
    else:
        raise ValueError(f"Unknown curve_source: {curve_source}")

    labels, bal, eq = curves_to_daily(curves)

    peak, low, max_dd = max_drawdown(eq)
    net = round(sum(t["profit"] + t["commission"] + t["swap"] for t in trades), 2)
    months = months_between(labels)
    max_open = max_open_positions(trades)

    print(f"  {curve_source}: {len(trades)} trades  |  net P&L: ${net:,.2f}  |  "
          f"peak: ${peak:,.2f}  |  max DD: ${max_dd:,.2f}  |  "
          f"max open: {max_open}  |  {months:.1f} months")

    return {
        "symbol": cfg["symbol"],
        "scale": scale,
        "base_lot": base_lot,
        "lot_size": round(base_lot * scale, 4),
        "trades": len(trades),
        "net": net,
        "peak": peak,
        "low": low,
        "max_dd": max_dd,
        "max_open_positions": max_open,
        "months": months,
        "labels": labels,
        "balance": bal,
        "equity": eq,
        "trade_windows": [(int(t["ts"]), int(t["close_ts"])) for t in trades],
    }


def combine_curves(strategies: List[Dict]) -> Dict:
    all_dates = sorted({d for s in strategies for d in s["labels"]})
    n = len(all_dates)
    combined_bal = [0.0] * n
    combined_eq = [0.0] * n

    for s in strategies:
        by_bal = dict(zip(s["labels"], s["balance"]))
        by_eq = dict(zip(s["labels"], s["equity"]))
        last_bal = 0.0
        last_eq = 0.0
        started = False
        for i, d in enumerate(all_dates):
            if d in by_bal:
                started = True
                last_bal = by_bal[d]
                last_eq = by_eq[d]
            if started:
                combined_bal[i] += last_bal
                combined_eq[i] += last_eq

    peak, low, max_dd = max_drawdown(combined_eq)
    return {
        "labels": all_dates,
        "balance": [round(x, 2) for x in combined_bal],
        "equity": [round(x, 2) for x in combined_eq],
        "net": round(combined_eq[-1] if combined_eq else 0, 2),
        "peak": peak,
        "low": low,
        "max_dd": max_dd,
        "max_open_positions": max_open_positions_for_strategies(strategies),
        "months": months_between(all_dates),
    }


def combine_curves_scaled(subset: List[Dict], scales: List[float]) -> Dict:
    """Like combine_curves but applies a per-strategy multiplier to
    each strategy's contribution. Used by the scaled optimizer to
    evaluate combinations with different lot multiples per strategy
    without having to build rescaled strategy dicts for every trial."""
    all_dates = sorted({d for s in subset for d in s["labels"]})
    n = len(all_dates)
    combined_bal = [0.0] * n
    combined_eq = [0.0] * n

    for s, scale in zip(subset, scales):
        by_bal = dict(zip(s["labels"], s["balance"]))
        by_eq = dict(zip(s["labels"], s["equity"]))
        last_bal = 0.0
        last_eq = 0.0
        started = False
        for i, d in enumerate(all_dates):
            if d in by_bal:
                started = True
                last_bal = by_bal[d]
                last_eq = by_eq[d]
            if started:
                combined_bal[i] += last_bal * scale
                combined_eq[i] += last_eq * scale

    peak, low, max_dd = max_drawdown(combined_eq)
    return {
        "labels": all_dates,
        "balance": [round(x, 2) for x in combined_bal],
        "equity": [round(x, 2) for x in combined_eq],
        "net": round(combined_eq[-1] if combined_eq else 0, 2),
        "peak": peak,
        "low": low,
        "max_dd": max_dd,
        "months": months_between(all_dates),
    }


def rescale_strategy(s: Dict, new_scale: float) -> Dict:
    """Return a copy of a strategy dict with all money values and
    lot sizes multiplied by (new_scale / old_scale). Used when we
    want to hand a scaled strategy off to write_portfolio_xlsx for
    a per-portfolio mini summary sheet."""
    ratio = new_scale / s["scale"] if s["scale"] else new_scale
    return {
        **s,
        "scale": new_scale,
        "lot_size": round(s["base_lot"] * new_scale, 4),
        "net": round(s["net"] * ratio, 2),
        "peak": s["peak"] * ratio,
        "low": s["low"] * ratio,
        "max_dd": s["max_dd"] * ratio,
        "balance": [round(b * ratio, 2) for b in s["balance"]],
        "equity": [round(e * ratio, 2) for e in s["equity"]],
    }


def find_optimal_combinations(
        strategies: List[Dict],
        backtest_months_override: Optional[float],
        account_size: float,
        dd_tolerance_pct: float,
        min_sf: float,
        min_monthly_pct: float,
        min_n: int = 1,
        max_n: int = 3,
        max_scale: int = 5) -> Tuple[List[Dict], int]:
    """Enumerate every non-empty subset of strategies with every
    integer scale combination from 1x to max_scale, evaluate the
    combined equity curve, and return the trials whose combined
    safety factor and monthly % meet both thresholds.

    Subset sizes are constrained to [min_n, max_n] to keep the
    search space tractable and match the user's "portfolio width"
    preference (e.g. "at most 3 strategies per portfolio").

    Returns (passing_results_sorted, total_combinations_tested).
    """
    from itertools import combinations as iter_combinations, product

    results: List[Dict] = []
    n = len(strategies)
    total = 0
    tested = 0

    lo = max(1, min_n)
    hi = min(n, max_n)
    max_scale = max(1, int(max_scale))

    subset_plans: List[Tuple[int, Tuple[int, ...], List[Dict]]] = []
    by_k_counts: Dict[int, int] = {}

    # Print strategy-to-index mapping for diagnostics
    print(f"  Strategy index mapping:")
    for i, s in enumerate(strategies):
        print(f"    [{i}] {s['symbol']}")

    for k in range(lo, hi + 1):
        valid_subsets = 0
        for idx_tuple in iter_combinations(range(n), k):
            subset = [strategies[i] for i in idx_tuple]
            symbols = [s["symbol"] for s in subset]
            if len(set(symbols)) != len(symbols):
                continue
            subset_plans.append((k, idx_tuple, subset))
            valid_subsets += 1
        by_k_counts[k] = valid_subsets
        total += valid_subsets * (max_scale ** k)

    print(f"  Search space: {total:,} total trials")
    for k in range(lo, hi + 1):
        valid_subsets = by_k_counts.get(k, 0)
        scale_trials = max_scale ** k
        print(f"    {k} strategies: {valid_subsets:,} valid subsets x "
              f"{scale_trials:,} scale tuples = {valid_subsets * scale_trials:,} trials")

    started_at = time.perf_counter()
    last_report_at = started_at
    report_every_trials = max(250, min(10_000, total // 200 if total else 250))
    best_pass: Optional[Dict] = None

    def print_progress(force: bool = False) -> None:
        nonlocal last_report_at
        now = time.perf_counter()
        if not force:
            if tested == 0:
                return
            if tested % report_every_trials != 0 and (now - last_report_at) < 3.0:
                return
        elapsed = now - started_at
        rate = (tested / elapsed) if elapsed > 0 else 0.0
        remaining = max(0, total - tested)
        eta = (remaining / rate) if rate > 0 else 0.0
        pct = (100.0 * tested / total) if total else 100.0
        print(f"  Progress: {tested:,}/{total:,} ({pct:5.1f}%)  |  "
              f"elapsed {format_elapsed(elapsed)}  |  "
              f"eta {format_elapsed(eta)}  |  "
              f"rate {rate:,.1f} trials/s  |  "
              f"passing {len(results):,}", flush=True)
        if best_pass is not None:
            parts = [f"{sym}x{scl}" for sym, scl in zip(best_pass["symbols"], best_pass["scales"])]
            sf_str = (f"{best_pass['safety_factor']:.2f}x"
                      if best_pass["safety_factor"] != float("inf") else "inf")
            print(f"    Best so far: {', '.join(parts)}  |  "
                  f"Monthly {best_pass['monthly_pct']:.2f}%  |  "
                  f"Safety {sf_str}  |  Net ${best_pass['net']:,.2f}  |  "
                  f"Max DD ${best_pass['max_dd']:,.2f}", flush=True)
        last_report_at = now

    current_k = None
    for k, idx_tuple, subset in subset_plans:
        if k != current_k:
            current_k = k
            print(f"  Testing {k}-strategy combinations...", flush=True)

        for scale_tuple in product(range(1, max_scale + 1), repeat=k):
            tested += 1
            combined = combine_curves_scaled(subset, list(scale_tuple))
            months = (backtest_months_override
                      if backtest_months_override is not None
                      else combined["months"])
            rm = compute_risk_metrics(
                combined["net"], combined["max_dd"], months,
                account_size, dd_tolerance_pct,
            )
            if (rm["safety_factor"] >= min_sf
                    and rm["monthly_pct"] >= min_monthly_pct):
                candidate = {
                    "strategy_indices": list(idx_tuple),
                    "symbols": [s["symbol"] for s in subset],
                    "scales": list(scale_tuple),
                    "lot_sizes": [
                        round(subset[j]["base_lot"] * scale_tuple[j], 4)
                        for j in range(k)
                    ],
                    "n": k,
                    "net": combined["net"],
                    "max_dd": combined["max_dd"],
                    "months": months,
                    "safety_factor": rm["safety_factor"],
                    "monthly_pct": rm["monthly_pct"],
                }
                results.append(candidate)

                if (best_pass is None
                        or (candidate["monthly_pct"], candidate["safety_factor"])
                        > (best_pass["monthly_pct"], best_pass["safety_factor"])):
                    best_pass = candidate
                    parts = [f"{sym}x{scl}" for sym, scl in zip(candidate["symbols"], candidate["scales"])]
                    sf_str = (f"{candidate['safety_factor']:.2f}x"
                              if candidate["safety_factor"] != float("inf") else "inf")
                    print(f"  New best pass at {tested:,}/{total:,}: "
                          f"{', '.join(parts)}  |  Monthly {candidate['monthly_pct']:.2f}%  |  "
                          f"Safety {sf_str}  |  Net ${candidate['net']:,.2f}  |  "
                          f"Max DD ${candidate['max_dd']:,.2f}", flush=True)

            print_progress()

    print_progress(force=True)
    elapsed = time.perf_counter() - started_at
    print(f"  Optimization complete: {tested:,} trials in {format_elapsed(elapsed)} "
          f"({(tested / elapsed) if elapsed > 0 else 0.0:,.1f} trials/s)", flush=True)

    # Rank: highest monthly % first, break ties by highest safety factor
    results.sort(key=lambda r: (-r["monthly_pct"], -r["safety_factor"]))
    return results, tested


def select_diverse_top_n(passing: List[Dict], n: int = 3) -> List[Dict]:
    """Greedy selection of N portfolios that use overlapping
    strategies as little as possible.

    Pick #1 = best passing combination (first element of the
    already-sorted list). For each subsequent pick, score all
    remaining candidates by:
      1. minimize number of strategies already used in prior picks
      2. among equal-overlap candidates, prefer higher monthly %
      3. tiebreak on higher safety factor
    """
    if not passing:
        return []

    # Safety filter: skip any result that somehow has duplicate symbols
    clean = [r for r in passing if len(set(r["symbols"])) == len(r["symbols"])]
    if not clean:
        return []

    picked = [clean[0]]
    remaining = list(clean[1:])

    while len(picked) < n and remaining:
        used_symbols: set = set()
        for p in picked:
            used_symbols.update(p["symbols"])

        best_idx = 0
        best_score: Optional[Tuple] = None
        for i, r in enumerate(remaining):
            overlap = len(set(r["symbols"]) & used_symbols)
            score = (-overlap, r["monthly_pct"], r["safety_factor"])
            if best_score is None or score > best_score:
                best_score = score
                best_idx = i
        picked.append(remaining.pop(best_idx))

    return picked


def build_optimization_text(results: List[Dict], total_combos: int,
                              min_sf: float, min_monthly_pct: float,
                              min_n: int = 1, max_n: int = 3,
                              max_scale: int = 5,
                              display_limit: int = 50) -> List[str]:
    lines = [""]
    lines.append("=" * 92)
    lines.append("OPTIMIZATION — Combinations meeting criteria")
    lines.append("=" * 92)
    lines.append("")
    lines.append(f"  Thresholds:")
    lines.append(f"    Safety Factor  >=  {min_sf:.2f}x")
    lines.append(f"    Monthly %      >=  {min_monthly_pct:.2f}%")
    lines.append(f"  Search space:")
    lines.append(f"    Strategies per combo:  {min_n} to {max_n}")
    lines.append(f"    Scale range:           1x to {max_scale}x per strategy")
    lines.append("")
    lines.append(f"  Tested:  {total_combos:,} trials")
    lines.append(f"  Passing: {len(results):,}")
    lines.append("")

    if not results:
        lines.append("  No combinations meet the criteria. Relax the thresholds")
        lines.append("  (lower MIN_SAFETY_FACTOR or MIN_MONTHLY_PCT) or widen")
        lines.append("  the search (raise MAX_STRATEGIES or MAX_SCALE) and rerun.")
        return lines

    lines.append(
        f"  {'Rank':>4} {'N':>3}  {'Strategies (x scale)':<46} "
        f"{'Net P&L':>12} {'Max DD':>12} {'Safety':>9} {'Monthly %':>11}"
    )
    lines.append(f"  {'-' * 90}")

    shown = results[:display_limit]
    for i, r in enumerate(shown, start=1):
        parts = [f"{sym}x{scl}" for sym, scl in zip(r["symbols"], r["scales"])]
        sym_str = ", ".join(parts)
        if len(sym_str) > 44:
            sym_str = sym_str[:41] + "..."
        sf_str = (f"{r['safety_factor']:.2f}x"
                  if r["safety_factor"] != float("inf") else "inf")
        lines.append(
            f"  {i:>4} {r['n']:>3}  {sym_str:<46} "
            f"${r['net']:>10,.2f} ${r['max_dd']:>10,.2f} "
            f"{sf_str:>9} {r['monthly_pct']:>10.2f}%"
        )

    if len(results) > display_limit:
        lines.append("")
        lines.append(f"  ... and {len(results) - display_limit:,} more. "
                     f"See the Optimization sheet in the xlsx for the full list.")

    lines.append("")
    lines.append("  Ranked by Monthly % descending, then Safety Factor descending.")
    return lines


def build_diverse_top_text(top: List[Dict]) -> List[str]:
    lines = [""]
    lines.append("=" * 92)
    lines.append("TOP DIVERSE PORTFOLIOS — greedy selection minimising symbol overlap")
    lines.append("=" * 92)
    lines.append("")

    if not top:
        lines.append("  No diverse portfolios to report — no combinations passed "
                     "the optimization criteria.")
        return lines

    for i, r in enumerate(top, start=1):
        parts = [f"{sym} x{scl}" for sym, scl in zip(r["symbols"], r["scales"])]
        sf_str = (f"{r['safety_factor']:.2f}x"
                  if r["safety_factor"] != float("inf") else "inf")
        lines.append(f"  #{i}  [{r['n']} strategies]  {', '.join(parts)}")
        lines.append(f"       Net ${r['net']:,.2f}  "
                     f"|  Max DD ${r['max_dd']:,.2f}  "
                     f"|  Safety {sf_str}  "
                     f"|  Monthly {r['monthly_pct']:.2f}%")
        lines.append("")

    lines.append("  Each top portfolio is saved to its own subfolder under the")
    lines.append("  output directory, containing a mini xlsx summary and the")
    lines.append("  backtest HTML + chart images for each strategy in that")
    lines.append("  portfolio (stem-globbed, so all sibling files are copied).")
    return lines


def build_stats_text(strategies: List[Dict], combined: Dict,
                      account_size: float, dd_tolerance_pct: float,
                      backtest_months_override: Optional[float]) -> List[str]:
    lines = [""]
    lines.append("=" * 92)
    lines.append("PORTFOLIO BACKTEST — Per-strategy + Combined")
    lines.append("=" * 92)
    lines.append("")
    lines.append(f"  {'Strategy':<14} {'Scale':>7} {'Trades':>8} {'Max Open':>9} "
                 f"{'Net P&L':>14} {'Peak':>14} {'Max DD':>14} {'Ret/DD':>10}")
    lines.append(f"  {'-' * 90}")

    for s in strategies:
        ret_dd = s["net"] / s["max_dd"] if s["max_dd"] > 0 else float("inf")
        ret_dd_str = f"{ret_dd:.2f}" if ret_dd != float("inf") else "inf"
        lines.append(
            f"  {s['symbol']:<14} {s['scale']:>6.2f}x {s['trades']:>8} {s['max_open_positions']:>9} "
            f"${s['net']:>12,.2f} ${s['peak']:>12,.2f} "
            f"${s['max_dd']:>12,.2f} {ret_dd_str:>10}"
        )

    lines.append(f"  {'-' * 90}")
    sum_dd_naive = sum(s["max_dd"] for s in strategies)
    combined_ret_dd = combined["net"] / combined["max_dd"] if combined["max_dd"] > 0 else float("inf")
    combined_ret_dd_str = f"{combined_ret_dd:.2f}" if combined_ret_dd != float("inf") else "inf"
    lines.append(
        f"  {'PORTFOLIO':<14} {'':>7} {'':>8} {combined['max_open_positions']:>9} "
        f"${combined['net']:>12,.2f} ${combined['peak']:>12,.2f} "
        f"${combined['max_dd']:>12,.2f} {combined_ret_dd_str:>10}"
    )
    lines.append("")
    lines.append(f"  Sum of individual max DDs (naive):  ${sum_dd_naive:,.2f}")
    lines.append(f"  Actual portfolio max DD:            ${combined['max_dd']:,.2f}")
    lines.append(f"  Max open positions at once:         {combined['max_open_positions']}")

    if sum_dd_naive > 0:
        saved = sum_dd_naive - combined["max_dd"]
        pct = 100 * saved / sum_dd_naive
        if saved > 0.5:
            lines.append(f"  Diversification benefit:            "
                         f"${saved:,.2f} ({pct:.1f}% lower than sum of parts)")
        else:
            lines.append(f"  Diversification benefit:            "
                         f"none — drawdowns aligned in time")

    # ── Risk metrics section ───────────────────────────────────────────
    allowable_dd = account_size * (dd_tolerance_pct / 100.0)
    lines.append("")
    lines.append("=" * 92)
    lines.append("RISK METRICS")
    lines.append("=" * 92)
    lines.append("")
    lines.append(f"  Account size:       ${account_size:,.2f}")
    lines.append(f"  DD tolerance:       {dd_tolerance_pct:.1f}%  "
                 f"(= ${allowable_dd:,.2f} allowable DD)")
    if backtest_months_override is not None:
        lines.append(f"  Backtest months:    {backtest_months_override:.1f} "
                     f"(override, applied to all rows)")
    else:
        lines.append(f"  Backtest months:    auto-computed per strategy "
                     f"from date range")
    lines.append("")
    lines.append(f"  {'Strategy':<14} {'Months':>8} {'Lot':>8} {'Open Pos':>9} "
                 f"{'Net P&L':>14} {'Max DD':>14} "
                 f"{'Safety Factor':>16} {'Monthly %':>12}")
    lines.append(f"  {'-' * 90}")

    for s in strategies:
        months = backtest_months_override if backtest_months_override is not None else s["months"]
        rm = compute_risk_metrics(
            s["net"], s["max_dd"], months, account_size, dd_tolerance_pct
        )
        sf_str = (f"{rm['safety_factor']:.2f}x"
                  if rm["safety_factor"] != float("inf") else "inf")
        lot_str = f"{s['lot_size']:.2f}"
        lines.append(
            f"  {s['symbol']:<14} {months:>7.1f}  {lot_str:>8} {s['max_open_positions']:>9} "
            f"${s['net']:>12,.2f} ${s['max_dd']:>12,.2f} "
            f"{sf_str:>16} {rm['monthly_pct']:>11.2f}%"
        )

    lines.append(f"  {'-' * 90}")
    p_months = (backtest_months_override
                if backtest_months_override is not None else combined["months"])
    p_rm = compute_risk_metrics(
        combined["net"], combined["max_dd"], p_months,
        account_size, dd_tolerance_pct
    )
    p_sf_str = (f"{p_rm['safety_factor']:.2f}x"
                if p_rm["safety_factor"] != float("inf") else "inf")
    lines.append(
        f"  {'PORTFOLIO':<14} {p_months:>7.1f}  {'':>8} {combined['max_open_positions']:>9} "
        f"${combined['net']:>12,.2f} ${combined['max_dd']:>12,.2f} "
        f"{p_sf_str:>16} {p_rm['monthly_pct']:>11.2f}%"
    )
    lines.append("")
    lines.append("  Safety Factor = (account × DD tolerance) / Max DD")
    lines.append("                  interpret as: multiples of current size")
    lines.append("                  you could trade before historical max DD")
    lines.append("                  would equal your tolerance limit.")
    lines.append("                  >1 = within budget, <1 = already exceeds.")
    lines.append("  Monthly %     = average monthly return as % of account.")
    lines.append("                  (Net P&L / Account) / Backtest Months")
    return lines


def write_portfolio_report(strategies, combined, stats_lines, out_path, title):
    datasets_eq = []
    datasets_bal = []
    for i, s in enumerate(strategies):
        color = STRATEGY_COLORS[i % len(STRATEGY_COLORS)]
        suffix = f" ({s['scale']:.2f}x)" if s["scale"] != 1.0 else ""
        by_eq = dict(zip(s["labels"], s["equity"]))
        by_bal = dict(zip(s["labels"], s["balance"]))
        aligned_eq, aligned_bal = [], []
        last_eq = 0.0
        last_bal = 0.0
        started = False
        for d in combined["labels"]:
            if d in by_eq:
                started = True
                last_eq = by_eq[d]
                last_bal = by_bal[d]
            aligned_eq.append(last_eq if started else None)
            aligned_bal.append(last_bal if started else None)
        datasets_eq.append({"label": s["symbol"] + suffix, "data": aligned_eq,
                            "color": color, "width": 1.5})
        datasets_bal.append({"label": s["symbol"] + suffix + " balance",
                             "data": aligned_bal, "color": color, "width": 1.5})

    datasets_eq.insert(0, {"label": "PORTFOLIO equity", "data": combined["equity"],
                           "color": COMBINED_COLOR, "width": 3})
    datasets_bal.insert(0, {"label": "PORTFOLIO balance", "data": combined["balance"],
                            "color": COMBINED_COLOR, "width": 3})

    def ds_json(datasets):
        return [{
            "label": d["label"], "data": d["data"],
            "borderColor": d["color"], "backgroundColor": d["color"],
            "borderWidth": d["width"], "fill": False,
            "pointRadius": 1.5, "pointHoverRadius": 5,
            "tension": 0, "spanGaps": True,
        } for d in datasets]

    stats_html = (
        '<pre style="font-family: Menlo, Consolas, monospace; font-size: 12px; '
        'background: #f8f8f8; padding: 16px; border-radius: 6px; '
        'white-space: pre; overflow-x: auto; border: 1px solid #e0e0e0;">'
        + html_lib.escape("\n".join(stats_lines))
        + "</pre>"
    )

    html_out = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>{html_lib.escape(title)}</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.js"></script>
<style>
  body {{ font-family: -apple-system, sans-serif; margin: 20px; background: #f7f7f7; color: #222; }}
  .container {{ max-width: 1300px; margin: auto; background: white; padding: 24px;
                border-radius: 8px; box-shadow: 0 2px 8px rgba(0,0,0,0.1); }}
  h1 {{ font-size: 22px; margin: 0 0 16px; }}
  h2 {{ font-size: 15px; margin: 24px 0 8px; color: #555; font-weight: 500; }}
  .chart-box {{ position: relative; height: 420px; margin-bottom: 24px; }}
</style></head><body><div class="container">
<h1>{html_lib.escape(title)}</h1>
<h2>Portfolio equity (per-strategy + combined)</h2>
<div class="chart-box"><canvas id="eq_chart"></canvas></div>
<h2>Portfolio balance (per-strategy + combined)</h2>
<div class="chart-box"><canvas id="bal_chart"></canvas></div>
{stats_html}
<script>
const labels = {json.dumps(combined["labels"])};
const eqDatasets = {json.dumps(ds_json(datasets_eq))};
const balDatasets = {json.dumps(ds_json(datasets_bal))};

function mkChart(canvasId, datasets) {{
  new Chart(document.getElementById(canvasId), {{
    type: 'line',
    data: {{ labels, datasets }},
    options: {{
      responsive: true, maintainAspectRatio: false,
      interaction: {{ mode: 'index', intersect: false }},
      plugins: {{ legend: {{ position: 'top' }},
        tooltip: {{ callbacks: {{ label: c => c.dataset.label + ': $' + (c.parsed.y || 0).toLocaleString() }} }} }},
      scales: {{
        x: {{ ticks: {{ maxTicksLimit: 20, maxRotation: 45 }} }},
        y: {{ ticks: {{ callback: v => '$' + v.toLocaleString() }} }}
      }}
    }}
  }});
}}

mkChart('eq_chart', eqDatasets);
mkChart('bal_chart', balDatasets);
</script></div></body></html>"""

    with open(out_path, "w", encoding="utf-8") as fh:
        fh.write(html_out)


# ────────────────────────────────────────────────────────────────────────────
# Main
# ────────────────────────────────────────────────────────────────────────────
def write_portfolio_xlsx(strategies, combined, out_path, title,
                          account_size, dd_tolerance_pct,
                          backtest_months_override,
                          opt_results=None, opt_total=0,
                          min_sf=1.5, min_monthly_pct=1.5):
    """Write an Excel workbook matching the Trading.xlsx pattern:
    input cells at the top (blue, editable), a per-strategy table
    with live formulas for Safety Factor and Monthly %, and a
    portfolio total row. Requires openpyxl (raises ImportError if
    unavailable, which main() catches and warns about)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Portfolio"

    # Styles
    blue_font = Font(name="Arial", size=10, color="0000FF", bold=False)
    black_font = Font(name="Arial", size=10, color="000000")
    bold_font = Font(name="Arial", size=10, bold=True)
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    title_font = Font(name="Arial", size=13, bold=True)

    header_fill = PatternFill("solid", start_color="305496")
    input_fill = PatternFill("solid", start_color="FFF2CC")
    total_fill = PatternFill("solid", start_color="D9E1F2")

    thin = Side(border_style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    # ── Title ────────────────────────────────────────────────────────
    ws["A1"] = title
    ws["A1"].font = title_font
    ws.merge_cells("A1:J1")

    # ── Inputs block (top) ───────────────────────────────────────────
    ws["A3"] = "Account Size"
    ws["A3"].font = bold_font
    ws["B3"] = account_size
    ws["B3"].font = blue_font
    ws["B3"].fill = input_fill
    ws["B3"].number_format = '"$"#,##0'
    ws["B3"].border = border

    ws["A4"] = "DD Tolerance"
    ws["A4"].font = bold_font
    ws["B4"] = dd_tolerance_pct / 100.0
    ws["B4"].font = blue_font
    ws["B4"].fill = input_fill
    ws["B4"].number_format = "0.0%"
    ws["B4"].border = border

    ws["A5"] = "Backtest Months"
    ws["A5"].font = bold_font
    if backtest_months_override is not None:
        ws["B5"] = backtest_months_override
        ws["B5"].font = blue_font
        ws["B5"].fill = input_fill
    else:
        ws["B5"] = "(auto per row)"
        ws["B5"].font = Font(name="Arial", size=9, italic=True, color="808080")
    ws["B5"].number_format = "0.0"
    ws["B5"].border = border

    ws["D3"] = "Blue = editable input. Yellow fill = tweak these and " \
                "formulas will recalculate."
    ws["D3"].font = Font(name="Arial", size=9, italic=True, color="808080")
    ws.merge_cells("D3:J3")

    # ── Table headers (row 7) ────────────────────────────────────────
    headers = [
        "Strategy", "Trades", "Max Open", "Lot Size", "Months",
        "Net P&L", "Max DD", "Allowable DD", "Safety Factor", "Monthly %",
    ]
    HEADER_ROW = 7
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=HEADER_ROW, column=col_idx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border

    # ── Per-strategy rows ────────────────────────────────────────────
    # Formulas use absolute references to the input cells so the user
    # can edit B3/B4/B5 and the whole sheet recalculates.
    ACCOUNT_REF = "$B$3"
    DDTOL_REF = "$B$4"
    MONTHS_OVERRIDE_REF = "$B$5"

    data_row = HEADER_ROW + 1
    first_data_row = data_row
    for s in strategies:
        months = (backtest_months_override if backtest_months_override is not None
                  else s["months"])

        ws.cell(row=data_row, column=1, value=s["symbol"]).font = black_font
        ws.cell(row=data_row, column=2, value=s["trades"]).number_format = "#,##0"
        ws.cell(row=data_row, column=3, value=s["max_open_positions"]).number_format = "0"
        ws.cell(row=data_row, column=4, value=s["lot_size"]).number_format = "0.00"

        # Months: blue (editable hardcode) since this is a per-row number
        # the user might want to override manually.
        months_cell = ws.cell(row=data_row, column=5, value=round(months, 1))
        months_cell.font = blue_font
        months_cell.fill = input_fill
        months_cell.number_format = "0.0"

        net_cell = ws.cell(row=data_row, column=6, value=round(s["net"], 2))
        net_cell.font = blue_font
        net_cell.fill = input_fill
        net_cell.number_format = '"$"#,##0.00;("$"#,##0.00);-'

        dd_cell = ws.cell(row=data_row, column=7, value=round(s["max_dd"], 2))
        dd_cell.font = blue_font
        dd_cell.fill = input_fill
        dd_cell.number_format = '"$"#,##0.00;("$"#,##0.00);-'

        # Allowable DD = account × DD tolerance
        ws.cell(row=data_row, column=8,
                value=f"={ACCOUNT_REF}*{DDTOL_REF}"
                ).number_format = '"$"#,##0.00'

        # Safety Factor = allowable / max DD  (handles DD=0)
        dd_ref = ws.cell(row=data_row, column=7).coordinate
        ws.cell(row=data_row, column=9,
                value=f'=IF(OR({dd_ref}=0,{dd_ref}=""),"",'
                      f'({ACCOUNT_REF}*{DDTOL_REF})/{dd_ref})'
                ).number_format = '0.00"x"'

        # Monthly % = (net / account) / months
        net_ref = ws.cell(row=data_row, column=6).coordinate
        months_ref = ws.cell(row=data_row, column=5).coordinate
        ws.cell(row=data_row, column=10,
                value=f'=IF(OR({ACCOUNT_REF}=0,{months_ref}=0,{months_ref}=""),"",'
                      f'({net_ref}/{ACCOUNT_REF})/{months_ref})'
                ).number_format = "0.00%"

        # Apply borders and formula font color
        for col_idx in range(1, 11):
            cell = ws.cell(row=data_row, column=col_idx)
            cell.border = border
            if col_idx in (8, 9, 10):  # formula cells
                cell.font = black_font

        data_row += 1

    last_data_row = data_row - 1

    # ── Portfolio total row ──────────────────────────────────────────
    p_months = (backtest_months_override if backtest_months_override is not None
                else combined["months"])

    ws.cell(row=data_row, column=1, value="PORTFOLIO").font = bold_font
    ws.cell(row=data_row, column=2, value="").alignment = center
    ws.cell(row=data_row, column=3, value=combined["max_open_positions"]).alignment = center
    ws.cell(row=data_row, column=4, value="").alignment = center

    p_months_cell = ws.cell(row=data_row, column=5, value=round(p_months, 1))
    p_months_cell.font = Font(name="Arial", size=10, bold=True, color="0000FF")
    p_months_cell.fill = input_fill
    p_months_cell.number_format = "0.0"

    p_net = ws.cell(row=data_row, column=6, value=round(combined["net"], 2))
    p_net.font = Font(name="Arial", size=10, bold=True, color="0000FF")
    p_net.fill = input_fill
    p_net.number_format = '"$"#,##0.00;("$"#,##0.00);-'

    p_dd = ws.cell(row=data_row, column=7, value=round(combined["max_dd"], 2))
    p_dd.font = Font(name="Arial", size=10, bold=True, color="0000FF")
    p_dd.fill = input_fill
    p_dd.number_format = '"$"#,##0.00;("$"#,##0.00);-'

    ws.cell(row=data_row, column=8,
            value=f"={ACCOUNT_REF}*{DDTOL_REF}"
            ).number_format = '"$"#,##0.00'
    p_dd_ref = ws.cell(row=data_row, column=7).coordinate
    ws.cell(row=data_row, column=9,
            value=f'=IF(OR({p_dd_ref}=0,{p_dd_ref}=""),"",'
                  f'({ACCOUNT_REF}*{DDTOL_REF})/{p_dd_ref})'
            ).number_format = '0.00"x"'
    p_net_ref = ws.cell(row=data_row, column=6).coordinate
    p_months_ref = ws.cell(row=data_row, column=5).coordinate
    ws.cell(row=data_row, column=10,
            value=f'=IF(OR({ACCOUNT_REF}=0,{p_months_ref}=0,{p_months_ref}=""),"",'
                  f'({p_net_ref}/{ACCOUNT_REF})/{p_months_ref})'
            ).number_format = "0.00%"

    for col_idx in range(1, 11):
        cell = ws.cell(row=data_row, column=col_idx)
        cell.fill = total_fill if col_idx not in (5, 6, 7) else input_fill
        cell.border = border
        if col_idx in (1, 8, 9, 10):
            cell.font = Font(name="Arial", size=10, bold=True)

    # ── Column widths ────────────────────────────────────────────────
    widths = {
        "A": 18, "B": 9, "C": 10, "D": 11, "E": 10,
        "F": 15, "G": 15, "H": 15, "I": 16, "J": 13,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[HEADER_ROW].height = 22

    # ── Optimization sheet (if results provided) ─────────────────────
    if opt_results is not None:
        opt_ws = wb.create_sheet("Optimization")

        opt_ws["A1"] = "Optimization Results"
        opt_ws["A1"].font = title_font
        opt_ws.merge_cells("A1:G1")

        # Criteria block
        opt_ws["A3"] = "Min Safety Factor"
        opt_ws["A3"].font = bold_font
        opt_ws["B3"] = min_sf
        opt_ws["B3"].font = blue_font
        opt_ws["B3"].fill = input_fill
        opt_ws["B3"].number_format = '0.00"x"'
        opt_ws["B3"].border = border

        opt_ws["A4"] = "Min Monthly %"
        opt_ws["A4"].font = bold_font
        opt_ws["B4"] = min_monthly_pct / 100.0
        opt_ws["B4"].font = blue_font
        opt_ws["B4"].fill = input_fill
        opt_ws["B4"].number_format = "0.00%"
        opt_ws["B4"].border = border

        opt_ws["A5"] = "Combinations tested"
        opt_ws["A5"].font = bold_font
        opt_ws["B5"] = opt_total
        opt_ws["B5"].font = black_font
        opt_ws["B5"].border = border

        opt_ws["A6"] = "Combinations passing"
        opt_ws["A6"].font = bold_font
        opt_ws["B6"] = len(opt_results)
        opt_ws["B6"].font = black_font
        opt_ws["B6"].border = border

        opt_ws["D3"] = ("Every non-empty subset of the input strategies was "
                        "tested. Rows below are the ones whose COMBINED safety "
                        "factor and monthly % both meet the criteria above.")
        opt_ws["D3"].font = Font(name="Arial", size=9, italic=True, color="808080")
        opt_ws["D3"].alignment = Alignment(wrap_text=True, vertical="top")
        opt_ws.merge_cells("D3:I5")

        # Results table header
        opt_headers = ["Rank", "N", "Strategies (x scale)", "Net P&L", "Max DD",
                       "Months", "Safety Factor", "Monthly %"]
        OPT_HEADER_ROW = 8
        for col_idx, h in enumerate(opt_headers, start=1):
            c = opt_ws.cell(row=OPT_HEADER_ROW, column=col_idx, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
            c.border = border

        # Data rows
        for i, r in enumerate(opt_results, start=1):
            row = OPT_HEADER_ROW + i
            opt_ws.cell(row=row, column=1, value=i).number_format = "0"
            opt_ws.cell(row=row, column=2, value=r["n"]).number_format = "0"

            # Strategies column shows "SYMBOL x SCALE" per entry
            if "scales" in r:
                sym_str = ", ".join(
                    f"{sym} x{scl}" for sym, scl
                    in zip(r["symbols"], r["scales"])
                )
            else:
                sym_str = ", ".join(r["symbols"])
            opt_ws.cell(row=row, column=3, value=sym_str)

            opt_ws.cell(row=row, column=4, value=round(r["net"], 2)
                        ).number_format = '"$"#,##0.00;("$"#,##0.00);-'
            opt_ws.cell(row=row, column=5, value=round(r["max_dd"], 2)
                        ).number_format = '"$"#,##0.00;("$"#,##0.00);-'
            opt_ws.cell(row=row, column=6, value=round(r["months"], 1)
                        ).number_format = "0.0"
            sf_val = (r["safety_factor"]
                      if r["safety_factor"] != float("inf") else 0)
            opt_ws.cell(row=row, column=7, value=round(sf_val, 4)
                        ).number_format = '0.00"x"'
            opt_ws.cell(row=row, column=8,
                        value=round(r["monthly_pct"] / 100.0, 6)
                        ).number_format = "0.00%"
            for col_idx in range(1, 9):
                cell = opt_ws.cell(row=row, column=col_idx)
                cell.border = border
                cell.font = black_font

        # Column widths
        opt_widths = {
            "A": 8, "B": 5, "C": 52, "D": 14, "E": 14,
            "F": 9, "G": 15, "H": 12,
        }
        for col, w in opt_widths.items():
            opt_ws.column_dimensions[col].width = w
        opt_ws.row_dimensions[1].height = 22
        opt_ws.row_dimensions[OPT_HEADER_ROW].height = 22

    wb.save(out_path)


def _sanitize_filename(s: str) -> str:
    """Replace characters that are invalid in Windows filenames."""
    bad = '<>:"/\\|?*'
    out = "".join(c if c not in bad else "_" for c in s)
    return out.strip() or "portfolio"


def export_top_portfolios(top: List[Dict], strategies: List[Dict],
                            configs: List[Dict], out_dir: str,
                            account_size: float, dd_tolerance_pct: float,
                            backtest_months_override: Optional[float]) -> List[str]:
    """Create one subfolder per top portfolio under `out_dir` and
    populate it with:

      - A mini summary.xlsx for that portfolio showing just its
        strategies at their chosen scales plus combined stats
      - The backtest file for each strategy, along with every
        sibling file (matched via stem glob — picks up the main
        .htm/.html plus all the .png chart images MT5 exports
        next to the main report)

    Returns the list of created subfolder paths.
    """
    import glob
    import shutil

    created: List[str] = []

    for i, r in enumerate(top, start=1):
        # Folder name includes rank + symbol/scale summary so the
        # user can tell portfolios apart without opening anything.
        label_parts = [f"{sym}-x{scl}" for sym, scl
                        in zip(r["symbols"], r["scales"])]
        folder_stem = f"TopPortfolio_{i}_" + "_".join(label_parts)
        folder_name = _sanitize_filename(folder_stem)
        # Cap length to keep under Windows MAX_PATH constraints
        if len(folder_name) > 100:
            folder_name = folder_name[:100]
        sub_dir = os.path.join(out_dir, folder_name)
        os.makedirs(sub_dir, exist_ok=True)
        created.append(sub_dir)

        print(f"\n  Top #{i}: {folder_name}")
        # Verify symbols match indices (catch any cached-vs-actual mismatch)
        actual_syms = [strategies[idx]["symbol"] for idx in r["strategy_indices"]]
        if actual_syms != r["symbols"]:
            print(f"    ⚠ Symbol mismatch! cached={r['symbols']}, "
                  f"actual={actual_syms}. Using actual.")
            r["symbols"] = actual_syms
        print(f"    Indices: {r['strategy_indices']}, "
              f"Symbols: {actual_syms}, Scales: {r['scales']}")

        # Build rescaled strategy list for the mini xlsx
        # Verify: no duplicate indices (should be impossible from combinations)
        if len(set(r["strategy_indices"])) != len(r["strategy_indices"]):
            print(f"    ⚠ SKIPPED — duplicate strategy indices: "
                  f"{r['strategy_indices']}")
            continue

        sub_strategies = [
            rescale_strategy(strategies[idx], scale)
            for idx, scale in zip(r["strategy_indices"], r["scales"])
        ]
        sub_combined = combine_curves(sub_strategies)

        sub_stats_lines = build_stats_text(
            sub_strategies, sub_combined,
            account_size, dd_tolerance_pct, backtest_months_override,
        )

        mini_xlsx_path = os.path.join(sub_dir, "portfolio_summary.xlsx")
        try:
            write_portfolio_xlsx(
                sub_strategies, sub_combined, mini_xlsx_path,
                f"Top Portfolio #{i}",
                account_size, dd_tolerance_pct, backtest_months_override,
                None, 0, 0, 0,  # no optimization sheet in mini
            )
            print(f"    ✓ {os.path.basename(mini_xlsx_path)}")
        except ImportError:
            print(f"    ⚠ openpyxl not installed — skipped xlsx summary")

        # Copy each strategy's backtest file + all stem-matching siblings
        copied_files: List[str] = []
        for idx in r["strategy_indices"]:
            cfg = configs[idx]
            bt_path = cfg["bt_path"]
            if not os.path.exists(bt_path):
                # The tool may have swapped .htm<->.html during loading;
                # try both variants to find the real file on disk.
                for alt in (bt_path[:-1] if bt_path.lower().endswith(".html")
                            else bt_path + "l"):
                    if os.path.exists(alt):
                        bt_path = alt
                        break
            if not os.path.exists(bt_path):
                print(f"    ⚠ backtest source missing: {bt_path}")
                continue

            src_dir = os.path.dirname(os.path.abspath(bt_path))
            stem = os.path.splitext(os.path.basename(bt_path))[0]
            pattern = os.path.join(src_dir, stem + "*")
            matches = [f for f in glob.glob(pattern) if os.path.isfile(f)]
            for f in matches:
                dest = os.path.join(sub_dir, os.path.basename(f))
                shutil.copy2(f, dest)
                copied_files.append(os.path.basename(f))

        if copied_files:
            print(f"    ✓ copied {len(copied_files)} backtest files")

    return created


def main() -> int:
    ap = argparse.ArgumentParser(
        description="Combine multiple fixed-lot backtests into a portfolio "
                    "equity view. Supports MT4 live, MT4 tester, MT5 tester.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    ap.add_argument("--strategy", action="append", required=True,
                    metavar="SYMBOL|BACKTEST|BARS|SCALE|GMT",
                    help="Strategy definition (repeat once per strategy). "
                         "Pipe-delimited. SCALE defaults to 1.0, GMT to 2.")
    ap.add_argument("--out-dir", default=".",
                    help="Output directory for the portfolio report.")
    ap.add_argument("--title", default="Portfolio Backtest",
                    help="Title for the HTML report.")
    ap.add_argument("--ticks-dir", default="",
                    help="Directory containing per-symbol tick CSV files (optional).")
    ap.add_argument("--tick-suffix", default="_GMT+2_US-DST.csv",
                    help="Suffix appended to each symbol to build tick filename.")
    ap.add_argument("--tick-gmt", type=int, default=2,
                    help="Timezone offset used to parse tick timestamps.")
    ap.add_argument("--curve-sources", default="auto",
                    help="Comma-separated sources to run: auto,bars,ticks. Default: auto")
    ap.add_argument("--account-size", type=float, default=10000.0,
                    metavar="AMOUNT",
                    help="Notional account size in $ used for risk metrics: "
                         "safety factor, monthly % estimate, and the xlsx "
                         "output. Default $10,000.")
    ap.add_argument("--dd-tolerance", type=float, default=10.0,
                    metavar="PERCENT",
                    help="Maximum allowable drawdown as a percentage of the "
                         "account (e.g. 10 for 10%%). Default 10%%. "
                         "Safety Factor = (account × tolerance) / max DD.")
    ap.add_argument("--backtest-months", type=float, default=None,
                    metavar="N",
                    help="Backtest duration in months, used for the monthly "
                         "% estimate. If omitted, computed automatically from "
                         "the date range of each strategy's curve.")
    ap.add_argument("--no-xlsx", action="store_true",
                    help="Skip the xlsx output and only write the HTML report.")
    ap.add_argument("--optimize", action="store_true",
                    help="Run subset optimization: evaluate every non-empty "
                         "combination of the provided strategies and list the "
                         "ones whose combined safety factor and monthly %% meet "
                         "the --min-safety-factor and --min-monthly-pct thresholds.")
    ap.add_argument("--min-safety-factor", type=float, default=1.5,
                    metavar="X",
                    help="Minimum safety factor for a combination to qualify "
                         "(default 1.5).")
    ap.add_argument("--min-monthly-pct", type=float, default=1.5,
                    metavar="PERCENT",
                    help="Minimum monthly %% return for a combination to "
                         "qualify (default 1.5).")
    ap.add_argument("--min-strategies", type=int, default=1, metavar="N",
                    help="Minimum number of strategies in each combination "
                         "tested by the optimizer (default 1).")
    ap.add_argument("--max-strategies", type=int, default=3, metavar="N",
                    help="Maximum number of strategies in each combination "
                         "tested by the optimizer (default 3). Keep this "
                         "small — search space grows quickly.")
    ap.add_argument("--max-scale", type=int, default=5, metavar="N",
                    help="Maximum integer scale multiplier tried per "
                         "strategy in the optimization search (default 5). "
                         "Each strategy is tested at scales 1, 2, ..., N.")
    ap.add_argument("--top-n", type=int, default=3, metavar="N",
                    help="After optimization, select this many top "
                         "portfolios that share strategies as little as "
                         "possible (default 3). Each gets its own subfolder "
                         "with a mini xlsx and copies of the backtest files.")
    args = ap.parse_args()

    os.makedirs(args.out_dir, exist_ok=True)

    try:
        configs = [parse_strategy_arg(s) for s in args.strategy]
        if args.ticks_dir:
            for cfg in configs:
                cfg["tick_path"] = os.path.join(args.ticks_dir, cfg["symbol"] + args.tick_suffix)
    except ValueError as e:
        print(f"ERROR: {e}", file=sys.stderr)
        return 1

    requested_sources = [s.strip().lower() for s in (args.curve_sources or "auto").split(",") if s.strip()]
    if not requested_sources:
        requested_sources = ["auto"]

    if "auto" in requested_sources:
        sources_to_run: List[str] = []
        has_bars = any((cfg.get("bars_path") or "").strip() and os.path.exists(cfg["bars_path"]) for cfg in configs)
        has_ticks = bool(args.ticks_dir)
        if has_bars:
            sources_to_run.append("bars")
        if has_ticks:
            sources_to_run.append("ticks")
        if not sources_to_run:
            print("ERROR: no usable bars or ticks source was supplied.", file=sys.stderr)
            return 1
    else:
        valid_sources = {"bars", "ticks"}
        invalid = [s for s in requested_sources if s not in valid_sources]
        if invalid:
            print(f"ERROR: invalid --curve-sources value(s): {', '.join(invalid)}. Use auto,bars,ticks", file=sys.stderr)
            return 1
        sources_to_run = []
        for src in requested_sources:
            if src not in sources_to_run:
                sources_to_run.append(src)

    if "ticks" in sources_to_run and not args.ticks_dir:
        print("ERROR: --ticks-dir is required when ticks source is enabled.", file=sys.stderr)
        return 1

    def run_source(source_name: str) -> None:
        print(f"\n{'=' * 92}")
        print(f"SOURCE RUN: {source_name.upper()}")
        print(f"{'=' * 92}")

        strategies = []
        for cfg in configs:
            try:
                strategies.append(load_strategy(cfg, args.backtest_months,
                                                curve_source=source_name,
                                                tick_gmt=args.tick_gmt))
            except (FileNotFoundError, ValueError) as e:
                print(f"\nERROR loading {cfg['symbol']} for {source_name}: {e}", file=sys.stderr)
                raise

        print("\nCombining curves on unified daily timeline…")
        combined = combine_curves(strategies)

        stats_lines = build_stats_text(
            strategies, combined,
            args.account_size, args.dd_tolerance, args.backtest_months
        )
        for ln in stats_lines:
            print(ln)

        opt_results: List[Dict] = []
        opt_total = 0
        if args.optimize:
            print(f"\nRunning optimization search "
                  f"(subset size {args.min_strategies}-{args.max_strategies}, "
                  f"scale 1-{args.max_scale}x)…")
            opt_results, opt_total = find_optimal_combinations(
                strategies, args.backtest_months,
                args.account_size, args.dd_tolerance,
                args.min_safety_factor, args.min_monthly_pct,
                args.min_strategies, args.max_strategies, args.max_scale,
            )
            opt_lines = build_optimization_text(
                opt_results, opt_total,
                args.min_safety_factor, args.min_monthly_pct,
                args.min_strategies, args.max_strategies, args.max_scale,
            )
            for ln in opt_lines:
                print(ln)
            stats_lines.extend(opt_lines)

            top_portfolios: List[Dict] = []
            if opt_results and args.top_n > 0:
                top_portfolios = select_diverse_top_n(opt_results, args.top_n)
                top_lines = build_diverse_top_text(top_portfolios)
                for ln in top_lines:
                    print(ln)
                stats_lines.extend(top_lines)
        else:
            top_portfolios = []

        source_out_dir = args.out_dir if len(sources_to_run) == 1 else os.path.join(args.out_dir, source_name)
        os.makedirs(source_out_dir, exist_ok=True)

        report_title = args.title if len(sources_to_run) == 1 else f"{args.title} ({source_name})"
        out_path = os.path.join(source_out_dir, "portfolio_report.html")
        write_portfolio_report(strategies, combined, stats_lines,
                               out_path, report_title)
        print(f"\n✓ {source_name} HTML report: {out_path}")

        if not args.no_xlsx:
            xlsx_path = os.path.join(source_out_dir, "portfolio_report.xlsx")
            try:
                write_portfolio_xlsx(
                    strategies, combined, xlsx_path, report_title,
                    args.account_size, args.dd_tolerance, args.backtest_months,
                    opt_results if args.optimize else None,
                    opt_total if args.optimize else 0,
                    args.min_safety_factor, args.min_monthly_pct,
                )
                print(f"✓ {source_name} xlsx report: {xlsx_path}")
            except ImportError:
                print("⚠ xlsx output skipped — openpyxl not installed. "
                      "Run: pip install openpyxl")

        if top_portfolios:
            print(f"\nExporting top {len(top_portfolios)} diverse portfolios for {source_name}…")
            export_top_portfolios(
                top_portfolios, strategies, configs, source_out_dir,
                args.account_size, args.dd_tolerance, args.backtest_months,
            )

    try:
        for source_name in sources_to_run:
            run_source(source_name)
    except (FileNotFoundError, ValueError):
        return 1

    return 0


if __name__ == "__main__":
    sys.exit(main())
