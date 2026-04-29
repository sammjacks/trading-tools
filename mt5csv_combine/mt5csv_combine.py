"""
mt5csv_combine.py — combine MT5 equity-curve CSV exports into a portfolio view.

Reads the tab-delimited, UTF-16 equity files exported from the MT5 strategy
tester or account history (columns: <DATE>  <BALANCE>  <EQUITY>  <DEPOSIT LOAD>)
and combines them on a unified daily timeline.  Produces the same HTML report
and xlsx spreadsheet as portfolio_backtest.py — no bars or tick files needed.

Fully standalone — only uses the Python standard library plus openpyxl for xlsx.

Usage:
    python mt5csv_combine.py ^
        --csv "EURUSD|path\\to\\EURUSD.csv" ^
        --csv "USDCAD|path\\to\\USDCAD.csv" ^
        --csv "USDJPY|path\\to\\USDJPY.csv|1.5" ^
        --out-dir results ^
        --title "Night-Scalper Portfolio" ^
        --account-size 10000 ^
        --dd-tolerance 10

--csv pipe-format:  LABEL|PATH[|SCALE]
  LABEL   display name in the report (e.g. symbol or strategy name)
  PATH    path to the MT5 equity CSV file
  SCALE   optional P&L multiplier (default 1.0)
"""

import argparse
import html as html_lib
import json
import os
import sys
from datetime import datetime
from typing import Dict, List, Optional, Tuple


# ─────────────────────────────────────────────────────────────────────────────
# Chart colours  (same palette as portfolio_backtest.py)
# ─────────────────────────────────────────────────────────────────────────────
STRATEGY_COLORS = [
    "#378ADD", "#E89611", "#2E9E5A", "#B83D8A", "#7C4DFF",
    "#C94444", "#00897B", "#F4B400", "#5C6BC0", "#26A69A",
    "#EF5350", "#8D6E63", "#AB47BC", "#66BB6A", "#FFA726",
]
COMBINED_COLOR = "#111111"


# ─────────────────────────────────────────────────────────────────────────────
# Maths / stats helpers
# ─────────────────────────────────────────────────────────────────────────────
def max_drawdown(values: List[float]) -> Tuple[float, float, float]:
    """Return (peak, low, max_dd).  Running peak starts at max(first, 0)."""
    if not values:
        return 0.0, 0.0, 0.0
    peak = max(values[0], 0.0)
    low = values[0]
    max_dd = 0.0
    running_peak = max(values[0], 0.0)
    for v in values:
        if v > running_peak:
            running_peak = v
        dd = running_peak - v
        if dd > max_dd:
            max_dd = dd
        if v > peak:
            peak = v
        if v < low:
            low = v
    return peak, low, max_dd


def max_drawdown_pct(values: List[float], baseline: float = 0.0) -> float:
    """Return max peak-to-trough drawdown as a percentage.

    `baseline` is the account size used to anchor the equity denominator
    (pass `account_size` so the pct is relative to a meaningful number).
    """
    if not values:
        return 0.0
    running_peak = max(values[0] + baseline, 0.0)
    max_dd_pct = 0.0
    for v in values:
        equity_value = v + baseline
        if equity_value > running_peak:
            running_peak = equity_value
        if running_peak > 0:
            dd_pct = (running_peak - equity_value) / running_peak * 100.0
            if dd_pct > max_dd_pct:
                max_dd_pct = dd_pct
    return max_dd_pct


def months_between(date_strs: List[str]) -> float:
    """Return elapsed months between first and last label.
    Accepts 'YYYY-MM-DD' or 'YYYY-MM-DD HH:MM' formats.
    Uses 30.44 days/month average.
    """
    if not date_strs or len(date_strs) < 2:
        return 0.0

    def _parse(s: str) -> datetime:
        for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                continue
        raise ValueError(f"Cannot parse date label: {s!r}")

    try:
        first = _parse(date_strs[0])
        last = _parse(date_strs[-1])
    except ValueError:
        return 0.0
    days = (last - first).days
    return days / 30.44 if days > 0 else 0.0


def compute_risk_metrics(
    net: float,
    max_dd: float,
    months: float,
    account_size: float,
    dd_tolerance_pct: float,
) -> Dict:
    """Compute safety factor and monthly % for fixed-lot strategies."""
    allowable_dd = account_size * (dd_tolerance_pct / 100.0)
    safety_factor = allowable_dd / max_dd if max_dd > 0 else float("inf")
    if months > 0 and account_size > 0:
        monthly_pct = (net / account_size) / months * 100.0
    else:
        monthly_pct = 0.0
    return {"allowable_dd": allowable_dd, "safety_factor": safety_factor,
            "monthly_pct": monthly_pct}


def _fit_label(text: str, width: int) -> str:
    s = str(text)
    if len(s) <= width:
        return s
    return s[:width - 3] + "..." if width > 3 else s[:width]


# ─────────────────────────────────────────────────────────────────────────────
# MT5 equity CSV reader
# ─────────────────────────────────────────────────────────────────────────────
def read_mt5_equity_csv(path: str) -> List[Tuple[str, float, float]]:
    """Parse an MT5 equity CSV file.

    Handles UTF-16 LE/BE BOM and UTF-8 BOM.
    Columns (any order): <DATE>, <BALANCE>, <EQUITY>, <DEPOSIT LOAD>.
    Angle brackets around header names are stripped automatically.

    Returns a list of (date_str, balance, equity) tuples where date_str
    is 'YYYY-MM-DD'.  Multiple intra-day rows are collapsed to the last
    value for each calendar day.
    """
    with open(path, "rb") as fh:
        raw = fh.read()

    if raw[:2] in (b"\xff\xfe", b"\xfe\xff"):
        text = raw.decode("utf-16")
    elif raw[:3] == b"\xef\xbb\xbf":
        text = raw[3:].decode("utf-8", errors="replace")
    elif raw[:200].count(b"\x00") > 50:
        # Heuristic: looks like UTF-16 without BOM
        try:
            text = raw.decode("utf-16")
        except UnicodeDecodeError:
            text = raw.decode("utf-8", errors="replace")
    else:
        text = raw.decode("utf-8", errors="replace")

    rows = text.splitlines()

    # Locate the header row
    col_date = col_balance = col_equity = None
    header_idx = None
    for i, row in enumerate(rows):
        cols = [c.strip().strip("<>").upper() for c in row.split("\t")]
        if "DATE" in cols and "BALANCE" in cols and "EQUITY" in cols:
            col_date = cols.index("DATE")
            col_balance = cols.index("BALANCE")
            col_equity = cols.index("EQUITY")
            header_idx = i
            break

    if header_idx is None:
        raise ValueError(
            f"Could not find DATE / BALANCE / EQUITY header in {path!r}.\n"
            "Expected tab-delimited MT5 equity CSV with angle-bracket column names."
        )

    # Parse data rows; collapse to daily (last row wins for each day)
    daily: Dict[str, Tuple[float, float]] = {}
    date_order: List[str] = []

    for row in rows[header_idx + 1:]:
        cols = row.split("\t")
        if len(cols) <= max(col_date, col_balance, col_equity):
            continue
        raw_date = cols[col_date].strip()
        raw_bal = cols[col_balance].strip()
        raw_eq = cols[col_equity].strip()
        if not raw_date or not raw_bal or not raw_eq:
            continue
        try:
            # Date format from MT5: "YYYY.MM.DD HH:MM" → normalise to "YYYY-MM-DD"
            day_part = raw_date.split(" ")[0]  # strip time if present
            day_str = day_part.replace(".", "-")
            # Validate it looks like a date
            datetime.strptime(day_str, "%Y-%m-%d")
            bal = float(raw_bal.replace(" ", "").replace(",", ""))
            eq = float(raw_eq.replace(" ", "").replace(",", ""))
        except (ValueError, IndexError):
            continue

        if day_str not in daily:
            date_order.append(day_str)
        daily[day_str] = (bal, eq)

    if not daily:
        raise ValueError(f"No data rows parsed from {path!r}")

    return [(d, daily[d][0], daily[d][1]) for d in date_order]


# ─────────────────────────────────────────────────────────────────────────────
# Strategy loading
# ─────────────────────────────────────────────────────────────────────────────
def load_strategy(label: str, path: str, scale: float = 1.0,
                  account_size: float = 0.0) -> Dict:
    """Load one MT5 equity CSV and build a strategy dict.

    The balance/equity values in the file are absolute (starting from the
    initial deposit).  We convert to relative P&L (initial → 0) so the
    curves are additive when combined across strategies.
    """
    print(f"  Loading: {label}  ({os.path.basename(path)})")
    rows = read_mt5_equity_csv(path)

    labels = [r[0] for r in rows]
    bal_abs = [r[1] for r in rows]
    eq_abs = [r[2] for r in rows]

    initial = bal_abs[0]  # initial deposit / starting balance

    # Convert to relative P&L and apply scale
    bal_rel = [round((b - initial) * scale, 2) for b in bal_abs]
    eq_rel = [round((e - initial) * scale, 2) for e in eq_abs]

    net = eq_rel[-1]
    peak, low, peak_dd = max_drawdown(eq_rel)
    dd_pct = max_drawdown_pct(eq_rel, baseline=account_size)
    months = months_between(labels)

    print(f"    {len(labels)} days  |  net P&L: ${net:,.2f}  |  "
          f"peak DD: ${peak_dd:,.2f}  |  {months:.1f} months")

    return {
        "symbol": label,
        "display_name": label,
        "scale": scale,
        "base_lot": 0.0,
        "lot_size": 0.0,
        "trades": 0,
        "net": net,
        "peak": peak,
        "low": low,
        "max_dd": peak_dd,
        "max_dd_pct": dd_pct,
        "max_open_positions": 0,
        "months": months,
        "risk_mode": "FIXED_LOT",
        "report_initial_deposit": initial,
        "labels": labels,
        "balance": bal_rel,
        "equity": eq_rel,
        "trade_windows": [],
    }


# ─────────────────────────────────────────────────────────────────────────────
# Curve combining  (same logic as portfolio_backtest.py)
# ─────────────────────────────────────────────────────────────────────────────
def combine_curves(strategies: List[Dict], account_size: float = 0.0) -> Dict:
    """Merge strategy P&L curves on a unified daily timeline using
    last-value forward-fill.  Each strategy starts contributing from its
    own first date."""
    if len(strategies) == 1:
        s = strategies[0]
        return {
            "labels": s["labels"],
            "balance": s["balance"],
            "equity": s["equity"],
            "net": s["net"],
            "peak": s["peak"],
            "low": s["low"],
            "max_dd": s["max_dd"],
            "max_dd_pct": s["max_dd_pct"],
            "max_open_positions": 0,
            "months": s["months"],
        }

    all_dates = sorted({d for s in strategies for d in s["labels"]})
    n = len(all_dates)
    combined_bal = [0.0] * n
    combined_eq = [0.0] * n

    for s in strategies:
        by_bal = dict(zip(s["labels"], s["balance"]))
        by_eq = dict(zip(s["labels"], s["equity"]))
        last_bal = 0.0
        last_eq = 0.0
        started = False
        for i, d in enumerate(all_dates):
            if d in by_bal:
                started = True
                last_bal = by_bal[d]
                last_eq = by_eq[d]
            if started:
                combined_bal[i] += last_bal
                combined_eq[i] += last_eq

    peak, low, peak_dd = max_drawdown(combined_eq)
    dd_pct = max_drawdown_pct(combined_eq, baseline=account_size)
    return {
        "labels": all_dates,
        "balance": [round(x, 2) for x in combined_bal],
        "equity": [round(x, 2) for x in combined_eq],
        "net": round(combined_eq[-1] if combined_eq else 0.0, 2),
        "peak": peak,
        "low": low,
        "max_dd": peak_dd,
        "max_dd_pct": dd_pct,
        "max_open_positions": 0,
        "months": months_between(all_dates),
    }


# ─────────────────────────────────────────────────────────────────────────────
# Stats text
# ─────────────────────────────────────────────────────────────────────────────
def build_stats_text(strategies: List[Dict], combined: Dict,
                     account_size: float, dd_tolerance_pct: float,
                     backtest_months_override: Optional[float]) -> List[str]:
    name_w = 40
    rule_w = 110

    lines = [""]
    lines.append("=" * rule_w)
    lines.append("PORTFOLIO (MT5 Equity CSV) — Per-strategy + Combined")
    lines.append("=" * rule_w)
    lines.append("")
    lines.append(f"  {'Strategy':<{name_w}} {'Scale':>7} {'Net P&L':>14} "
                 f"{'Peak':>14} {'Max DD':>14} {'Ret/DD':>10}")
    lines.append(f"  {'-' * (rule_w - 2)}")

    for s in strategies:
        ret_dd = s["net"] / s["max_dd"] if s["max_dd"] > 0 else float("inf")
        ret_dd_str = f"{ret_dd:.2f}" if ret_dd != float("inf") else "inf"
        sym = _fit_label(s.get("display_name", s["symbol"]), name_w)
        lines.append(
            f"  {sym:<{name_w}} {s['scale']:>6.2f}x "
            f"${s['net']:>12,.2f} ${s['peak']:>12,.2f} "
            f"${s['max_dd']:>12,.2f} {ret_dd_str:>10}"
        )

    lines.append(f"  {'-' * (rule_w - 2)}")
    sum_dd_naive = sum(s["max_dd"] for s in strategies)
    comb_ret_dd = combined["net"] / combined["max_dd"] if combined["max_dd"] > 0 else float("inf")
    comb_ret_dd_str = f"{comb_ret_dd:.2f}" if comb_ret_dd != float("inf") else "inf"
    lines.append(
        f"  {'PORTFOLIO':<{name_w}} {'':>7} "
        f"${combined['net']:>12,.2f} ${combined['peak']:>12,.2f} "
        f"${combined['max_dd']:>12,.2f} {comb_ret_dd_str:>10}"
    )
    lines.append("")
    lines.append(f"  Sum of individual max DDs (naive):  ${sum_dd_naive:,.2f}")
    lines.append(f"  Actual portfolio max DD:            ${combined['max_dd']:,.2f}")
    if sum_dd_naive > 0:
        saved = sum_dd_naive - combined["max_dd"]
        pct = 100.0 * saved / sum_dd_naive
        if saved > 0.5:
            lines.append(f"  Diversification benefit:            "
                         f"${saved:,.2f} ({pct:.1f}% lower than sum of parts)")
        else:
            lines.append(f"  Diversification benefit:            "
                         f"none — drawdowns aligned in time")

    allowable_dd = account_size * (dd_tolerance_pct / 100.0)
    lines.append("")
    lines.append("=" * rule_w)
    lines.append("RISK METRICS")
    lines.append("=" * rule_w)
    lines.append("")
    lines.append(f"  Account size:       ${account_size:,.2f}")
    lines.append(f"  DD tolerance:       {dd_tolerance_pct:.1f}%"
                 f"  (= ${allowable_dd:,.2f} allowable DD)")
    if backtest_months_override is not None:
        lines.append(f"  Backtest months:    {backtest_months_override:.1f} (override)")
    else:
        lines.append(f"  Backtest months:    auto-computed per strategy from date range")
    lines.append("")
    lines.append(f"  {'Strategy':<{name_w}} {'Months':>8} {'Net P&L':>14} "
                 f"{'Max DD':>14} {'Safety Factor':>16} {'Monthly %':>12}")
    lines.append(f"  {'-' * (rule_w - 2)}")

    for s in strategies:
        months = backtest_months_override if backtest_months_override is not None else s["months"]
        rm = compute_risk_metrics(s["net"], s["max_dd"], months, account_size, dd_tolerance_pct)
        sf_str = f"{rm['safety_factor']:.2f}x" if rm["safety_factor"] != float("inf") else "inf"
        sym = _fit_label(s.get("display_name", s["symbol"]), name_w)
        lines.append(
            f"  {sym:<{name_w}} {months:>7.1f}  "
            f"${s['net']:>12,.2f} ${s['max_dd']:>12,.2f} "
            f"{sf_str:>16} {rm['monthly_pct']:>11.2f}%"
        )

    lines.append(f"  {'-' * (rule_w - 2)}")
    p_months = backtest_months_override if backtest_months_override is not None else combined["months"]
    p_rm = compute_risk_metrics(combined["net"], combined["max_dd"], p_months,
                                account_size, dd_tolerance_pct)
    p_sf_str = f"{p_rm['safety_factor']:.2f}x" if p_rm["safety_factor"] != float("inf") else "inf"
    lines.append(
        f"  {'PORTFOLIO':<{name_w}} {p_months:>7.1f}  "
        f"${combined['net']:>12,.2f} ${combined['max_dd']:>12,.2f} "
        f"{p_sf_str:>16} {p_rm['monthly_pct']:>11.2f}%"
    )
    lines.append("")
    lines.append("  Safety Factor = (account × DD tolerance) / Max DD  "
                 ">1 = within budget, <1 = exceeds budget")
    lines.append("  Monthly %     = (Net P&L / Account) / Months  "
                 "(linear average monthly return)")
    return lines


# ─────────────────────────────────────────────────────────────────────────────
# HTML report  (same template as portfolio_backtest.py)
# ─────────────────────────────────────────────────────────────────────────────
def write_portfolio_report(strategies: List[Dict], combined: Dict,
                           stats_lines: List[str], out_path: str,
                           title: str) -> None:
    datasets_eq: List[Dict] = []
    datasets_bal: List[Dict] = []

    for i, s in enumerate(strategies):
        color = STRATEGY_COLORS[i % len(STRATEGY_COLORS)]
        base_label = s.get("display_name", s["symbol"])
        suffix = f" ({s['scale']:.2f}x)" if s["scale"] != 1.0 else ""
        by_eq = dict(zip(s["labels"], s["equity"]))
        by_bal = dict(zip(s["labels"], s["balance"]))
        aligned_eq: List = []
        aligned_bal: List = []
        last_eq = 0.0
        last_bal = 0.0
        started = False
        for d in combined["labels"]:
            if d in by_eq:
                started = True
                last_eq = by_eq[d]
                last_bal = by_bal[d]
            aligned_eq.append(last_eq if started else None)
            aligned_bal.append(last_bal if started else None)
        datasets_eq.append({"label": base_label + suffix,
                             "data": aligned_eq, "color": color, "width": 1.5})
        datasets_bal.append({"label": base_label + suffix + " balance",
                              "data": aligned_bal, "color": color, "width": 1.5})

    datasets_eq.insert(0, {"label": "PORTFOLIO equity",
                            "data": combined["equity"],
                            "color": COMBINED_COLOR, "width": 3})
    datasets_bal.insert(0, {"label": "PORTFOLIO balance",
                             "data": combined["balance"],
                             "color": COMBINED_COLOR, "width": 3})

    def _ds_json(datasets: List[Dict]) -> List[Dict]:
        return [{
            "label": d["label"], "data": d["data"],
            "borderColor": d["color"], "backgroundColor": d["color"],
            "borderWidth": d["width"], "fill": False,
            "pointRadius": 1.5, "pointHoverRadius": 5,
            "tension": 0, "spanGaps": True,
        } for d in datasets]

    stats_html = (
        '<pre style="font-family: Menlo, Consolas, monospace; font-size: 12px; '
        'background: #f8f8f8; padding: 16px; border-radius: 6px; '
        'white-space: pre; overflow-x: auto; border: 1px solid #e0e0e0;">'
        + html_lib.escape("\n".join(stats_lines))
        + "</pre>"
    )

    html_out = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>{html_lib.escape(title)}</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.js"></script>
<style>
  body {{ font-family: -apple-system, sans-serif; margin: 20px; background: #f7f7f7; color: #222; }}
  .container {{ max-width: 1300px; margin: auto; background: white; padding: 24px;
                border-radius: 8px; box-shadow: 0 2px 8px rgba(0,0,0,0.1); }}
  h1 {{ font-size: 22px; margin: 0 0 16px; }}
  h2 {{ font-size: 15px; margin: 24px 0 8px; color: #555; font-weight: 500; }}
  .chart-box {{ position: relative; height: 420px; margin-bottom: 24px; }}
</style></head><body><div class="container">
<h1>{html_lib.escape(title)}</h1>
<h2>Portfolio equity (per-strategy + combined)</h2>
<div class="chart-box"><canvas id="eq_chart"></canvas></div>
<h2>Portfolio balance (per-strategy + combined)</h2>
<div class="chart-box"><canvas id="bal_chart"></canvas></div>
{stats_html}
<script>
const labels = {json.dumps(combined["labels"])};
const eqDatasets = {json.dumps(_ds_json(datasets_eq))};
const balDatasets = {json.dumps(_ds_json(datasets_bal))};

function mkChart(canvasId, datasets) {{
  new Chart(document.getElementById(canvasId), {{
    type: 'line',
    data: {{ labels, datasets }},
    options: {{
      responsive: true, maintainAspectRatio: false,
      interaction: {{ mode: 'index', intersect: false }},
      plugins: {{ legend: {{ position: 'top' }},
        tooltip: {{ callbacks: {{ label: c => c.dataset.label + ': $' +
          (c.parsed.y == null ? 'n/a' : c.parsed.y.toLocaleString(undefined,
            {{minimumFractionDigits:2,maximumFractionDigits:2}})) }} }} }},
      scales: {{
        x: {{ ticks: {{ maxTicksLimit: 20, maxRotation: 45 }} }},
        y: {{ ticks: {{ callback: v => '$' + v.toLocaleString() }} }}
      }}
    }}
  }});
}}

mkChart('eq_chart', eqDatasets);
mkChart('bal_chart', balDatasets);
</script></div></body></html>"""

    with open(out_path, "w", encoding="utf-8") as fh:
        fh.write(html_out)


# ─────────────────────────────────────────────────────────────────────────────
# xlsx report
# ─────────────────────────────────────────────────────────────────────────────
def write_portfolio_xlsx(strategies: List[Dict], combined: Dict,
                         out_path: str, title: str,
                         account_size: float, dd_tolerance_pct: float,
                         backtest_months_override: Optional[float]) -> None:
    """Write a Trading.xlsx-style workbook.  Requires openpyxl."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Portfolio"

    # ── Shared styles ────────────────────────────────────────────────
    blue_font = Font(name="Arial", size=10, color="0000FF")
    black_font = Font(name="Arial", size=10, color="000000")
    bold_font = Font(name="Arial", size=10, bold=True)
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    title_font = Font(name="Arial", size=13, bold=True)

    header_fill = PatternFill("solid", start_color="305496")
    input_fill = PatternFill("solid", start_color="FFF2CC")
    total_fill = PatternFill("solid", start_color="D9E1F2")

    thin = Side(border_style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    # ── Title ────────────────────────────────────────────────────────
    ws["A1"] = title
    ws["A1"].font = title_font
    ws.merge_cells("A1:H1")

    # ── Input cells ──────────────────────────────────────────────────
    ws["A3"] = "Account Size"
    ws["A3"].font = bold_font
    ws["B3"] = account_size
    ws["B3"].font = blue_font
    ws["B3"].fill = input_fill
    ws["B3"].number_format = '"$"#,##0'
    ws["B3"].border = border

    ws["A4"] = "DD Tolerance"
    ws["A4"].font = bold_font
    ws["B4"] = dd_tolerance_pct / 100.0
    ws["B4"].font = blue_font
    ws["B4"].fill = input_fill
    ws["B4"].number_format = "0.0%"
    ws["B4"].border = border

    ws["A5"] = "Backtest Months"
    ws["A5"].font = bold_font
    if backtest_months_override is not None:
        ws["B5"] = backtest_months_override
        ws["B5"].font = blue_font
        ws["B5"].fill = input_fill
    else:
        ws["B5"] = "(auto per row)"
        ws["B5"].font = Font(name="Arial", size=9, italic=True, color="808080")
    ws["B5"].number_format = "0.0"
    ws["B5"].border = border

    ws["D3"] = "Blue = editable input.  Yellow fill = tweak these and formulas recalculate."
    ws["D3"].font = Font(name="Arial", size=9, italic=True, color="808080")
    ws.merge_cells("D3:H3")

    # ── Table header (row 7) ─────────────────────────────────────────
    HEADER_ROW = 7
    headers = ["Strategy", "Scale", "Months", "Net P&L",
               "Max DD", "Allowable DD", "Safety Factor", "Monthly %"]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=HEADER_ROW, column=col_idx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border

    ACCOUNT_REF = "$B$3"
    DDTOL_REF = "$B$4"

    # ── Per-strategy rows ────────────────────────────────────────────
    data_row = HEADER_ROW + 1
    first_data_row = data_row

    for s in strategies:
        months = (backtest_months_override if backtest_months_override is not None
                  else s["months"])

        ws.cell(row=data_row, column=1,
                value=s.get("display_name", s["symbol"])).font = black_font

        sc = ws.cell(row=data_row, column=2, value=s["scale"])
        sc.font = black_font
        sc.number_format = '0.00"x"'

        mc = ws.cell(row=data_row, column=3, value=float(months))
        mc.font = blue_font
        mc.fill = input_fill
        mc.number_format = "0.0"

        nc = ws.cell(row=data_row, column=4, value=round(s["net"], 2))
        nc.font = blue_font
        nc.fill = input_fill
        nc.number_format = '"$"#,##0.00;("$"#,##0.00);-'

        dc = ws.cell(row=data_row, column=5, value=round(s["max_dd"], 2))
        dc.font = blue_font
        dc.fill = input_fill
        dc.number_format = '"$"#,##0.00;("$"#,##0.00);-'

        # Allowable DD formula
        ws.cell(row=data_row, column=6,
                value=f"={ACCOUNT_REF}*{DDTOL_REF}"
                ).number_format = '"$"#,##0.00'

        # Safety Factor formula
        dd_ref = f"E{data_row}"
        allow_ref = f"F{data_row}"
        ws.cell(row=data_row, column=7,
                value=f'=IF(OR({dd_ref}=0,{dd_ref}=""),"",'
                      f'{allow_ref}/{dd_ref})'
                ).number_format = '0.00"x"'

        # Monthly % formula
        net_ref = f"D{data_row}"
        months_ref = f"C{data_row}"
        ws.cell(row=data_row, column=8,
                value=f'=IF(OR({ACCOUNT_REF}=0,{months_ref}=0,{months_ref}=""),"",'
                      f'({net_ref}/{ACCOUNT_REF})/{months_ref})'
                ).number_format = "0.00%"

        for col_idx in range(1, 9):
            cell = ws.cell(row=data_row, column=col_idx)
            cell.border = border
            if col_idx in (6, 7, 8):
                cell.font = black_font

        data_row += 1

    last_data_row = data_row - 1

    # ── Portfolio total row ──────────────────────────────────────────
    p_months = (backtest_months_override if backtest_months_override is not None
                else combined["months"])

    ws.cell(row=data_row, column=1, value="PORTFOLIO").font = bold_font
    ws.cell(row=data_row, column=2, value="")

    pm = ws.cell(row=data_row, column=3, value=float(p_months))
    pm.font = Font(name="Arial", size=10, bold=True, color="0000FF")
    pm.fill = input_fill
    pm.number_format = "0.0"

    pn = ws.cell(row=data_row, column=4,
                 value=f"=SUM(D{first_data_row}:D{last_data_row})")
    pn.font = Font(name="Arial", size=10, bold=True, color="0000FF")
    pn.fill = input_fill
    pn.number_format = '"$"#,##0.00;("$"#,##0.00);-'

    pd = ws.cell(row=data_row, column=5, value=round(combined["max_dd"], 2))
    pd.font = Font(name="Arial", size=10, bold=True, color="0000FF")
    pd.fill = input_fill
    pd.number_format = '"$"#,##0.00;("$"#,##0.00);-'

    ws.cell(row=data_row, column=6,
            value=f"={ACCOUNT_REF}*{DDTOL_REF}"
            ).number_format = '"$"#,##0.00'

    p_dd_ref = f"E{data_row}"
    p_allow_ref = f"F{data_row}"
    ws.cell(row=data_row, column=7,
            value=f'=IF(OR({p_dd_ref}=0,{p_dd_ref}=""),"",'
                  f'{p_allow_ref}/{p_dd_ref})'
            ).number_format = '0.00"x"'

    p_net_ref = f"D{data_row}"
    p_months_ref = f"C{data_row}"
    ws.cell(row=data_row, column=8,
            value=f'=IF(OR({ACCOUNT_REF}=0,{p_months_ref}=0,{p_months_ref}=""),"",'
                  f'({p_net_ref}/{ACCOUNT_REF})/{p_months_ref})'
            ).number_format = "0.00%"

    for col_idx in range(1, 9):
        cell = ws.cell(row=data_row, column=col_idx)
        cell.fill = total_fill if col_idx not in (3, 4, 5) else input_fill
        cell.border = border
        if col_idx in (1, 6, 7, 8):
            cell.font = Font(name="Arial", size=10, bold=True)

    # ── Column widths ────────────────────────────────────────────────
    widths = {"A": 42, "B": 10, "C": 10, "D": 15,
              "E": 15, "F": 15, "G": 16, "H": 13}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[HEADER_ROW].height = 22

    wb.save(out_path)


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────
def _parse_csv_arg(s: str) -> Tuple[str, str, float]:
    """Parse 'LABEL|PATH[|SCALE]' → (label, path, scale)."""
    parts = [p.strip() for p in s.split("|")]
    if len(parts) < 2:
        raise ValueError(
            f"--csv value must be 'LABEL|PATH[|SCALE]', got: {s!r}"
        )
    label = parts[0]
    path = parts[1]
    scale = float(parts[2]) if len(parts) >= 3 and parts[2] else 1.0
    return label, path, scale


def main() -> int:
    ap = argparse.ArgumentParser(
        description="Combine MT5 equity CSV exports into a portfolio report.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    ap.add_argument("--csv", action="append", required=True,
                    metavar="LABEL|PATH[|SCALE]",
                    help="MT5 equity CSV file.  Repeat once per strategy.")
    ap.add_argument("--out-dir", default=".",
                    help="Output directory (created if absent).")
    ap.add_argument("--title", default="Portfolio — MT5 Equity Curves",
                    help="Title shown in the HTML report and xlsx.")
    ap.add_argument("--account-size", type=float, default=10000.0,
                    metavar="AMOUNT",
                    help="Notional account size in $ for risk metrics. "
                         "Default $10,000.")
    ap.add_argument("--dd-tolerance", type=float, default=10.0,
                    metavar="PERCENT",
                    help="Max allowable drawdown as %% of account. "
                         "Default 10%%.")
    ap.add_argument("--backtest-months", type=float, default=None,
                    metavar="N",
                    help="Override backtest duration in months "
                         "(default: auto-computed from each file's date range).")
    ap.add_argument("--no-xlsx", action="store_true",
                    help="Skip xlsx output.")
    args = ap.parse_args()

    os.makedirs(args.out_dir, exist_ok=True)

    strategies: List[Dict] = []
    for csv_arg in args.csv:
        try:
            label, path, scale = _parse_csv_arg(csv_arg)
        except ValueError as e:
            print(f"ERROR: {e}", file=sys.stderr)
            return 1
        if not os.path.exists(path):
            print(f"ERROR: file not found: {path!r}", file=sys.stderr)
            return 1
        try:
            strategies.append(
                load_strategy(label, path, scale, args.account_size)
            )
        except (ValueError, OSError) as e:
            print(f"ERROR loading {label!r}: {e}", file=sys.stderr)
            return 1

    print("\nCombining curves on unified daily timeline…")
    combined = combine_curves(strategies, account_size=args.account_size)

    stats_lines = build_stats_text(
        strategies, combined,
        args.account_size, args.dd_tolerance, args.backtest_months,
    )
    for ln in stats_lines:
        print(ln)

    html_path = os.path.join(args.out_dir, "portfolio_report.html")
    write_portfolio_report(strategies, combined, stats_lines,
                           html_path, args.title)
    print(f"\n✓ HTML report:  {html_path}")

    if not args.no_xlsx:
        xlsx_path = os.path.join(args.out_dir, "portfolio_report.xlsx")
        try:
            write_portfolio_xlsx(
                strategies, combined, xlsx_path, args.title,
                args.account_size, args.dd_tolerance, args.backtest_months,
            )
            print(f"✓ xlsx report:  {xlsx_path}")
        except ImportError:
            print("⚠  xlsx output skipped — openpyxl not installed.  "
                  "Run: pip install openpyxl")

    return 0


if __name__ == "__main__":
    sys.exit(main())
